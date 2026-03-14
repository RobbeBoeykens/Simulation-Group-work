import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DOWNTIME_COSTS = list(range(0, 2501, 10))
REPLACE_COST = 500

def run_simulation(failure_probs, warmup_periods, number_periods, runs, downtime_cost, seed=42):
    rng = random.Random(seed)
    total_periods = warmup_periods + number_periods
    average_costs_warmup = []

    for _ in range(runs):
        current_state = 0
        post_warmup_cost = 0

        for period in range(total_periods):
            p = failure_probs[current_state]

            if p == "replace":
                cost = REPLACE_COST
                current_state = 0
            else:
                r = rng.random()
                if p is None or r < p:
                    cost = downtime_cost + REPLACE_COST
                    current_state = 0
                else:
                    cost = 0
                    current_state += 1

            if period >= warmup_periods:
                post_warmup_cost += cost

        average_costs_warmup.append(post_warmup_cost / number_periods)

    return sum(average_costs_warmup) / runs


def machine_simulation_sensitivity(
    number_periods=10000,
    runs=100,
    filename="output_sensitivity.xlsx"
):
    policies = {
        "policy_0": {"failure_probs": [0.1, 0.2, 0.5, None],      "warmup": 1000},
        "policy_1": {"failure_probs": [0.1, 0.2, 0.5, "replace"],  "warmup": 1500},
        "policy_2": {"failure_probs": [0.1, 0.2, "replace"],       "warmup": 1500},
        "policy_3": {"failure_probs": [0.1, "replace"],            "warmup": 1500},
    }

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"


    header_font = Font(bold=True)
    best_fill   = PatternFill("solid", start_color="C6EFCE")
    header_fill = PatternFill("solid", start_color="D9D9D9")

    ws_summary.append(["Downtime cost"] + list(policies.keys()) + ["Best policy"])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill

    results = {}
    total_cost_levels = len(DOWNTIME_COSTS)

    for i, downtime_cost in enumerate(DOWNTIME_COSTS, start=1):
        progress = (i / total_cost_levels) * 100
        print(f"Running sensitivity {i}/{total_cost_levels} (downtime_cost={downtime_cost}) - {progress:.1f}% done", flush=True)
        policy_avgs = {}

        for policy_name, policy_data in policies.items():
            avg = run_simulation(
                failure_probs=policy_data["failure_probs"],
                warmup_periods=policy_data["warmup"],
                number_periods=number_periods,
                runs=runs,
                downtime_cost=downtime_cost,
            )
            policy_avgs[policy_name] = avg

        best = min(policy_avgs, key=policy_avgs.get)
        results[downtime_cost] = policy_avgs

        row_data = [downtime_cost] + [round(v, 4) for v in policy_avgs.values()] + [best]
        ws_summary.append(row_data)

        best_col = list(policies.keys()).index(best) + 2
        ws_summary.cell(row=ws_summary.max_row, column=best_col).fill = best_fill


    for col in ws_summary.columns:
        ws_summary.column_dimensions[col[0].column_letter].width = 18

    wb.save(filename)
    print(f"\nResultaten opgeslagen in {filename}")


machine_simulation_sensitivity(number_periods=10000, runs=100)