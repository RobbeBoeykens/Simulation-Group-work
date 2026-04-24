
import math
import os
import random

import openpyxl
from openpyxl import load_workbook

from simulation import Simulation

print(os.getcwd())

CONFIGURATIONS = [
    [12, 3, 1],
    [16, 1, 2],
    [14, 2, 4],
    [10, 2, 3],
    [20, 2, 1],
    [16, 3, 3],
    [14, 3, 2],
    [14, 1, 1],
    [10, 1, 2],
    [20, 1, 4],
]

WARMUP_WEEKS: int = 50     
RUN_WEEKS:    int = 483     
R:            int = 30     

EXCEL_PATH:   str = "Excel Files/replications.xlsx"
INPUT_DIR:    str = "Big Assignment/Inputs"

def safe_avg(values: list[float]) -> float:
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def sheet_name_for(config: list[int]) -> str:
    n_slots, strategy, rule = config
    return f"S{strategy}-{n_slots}slots-R{rule}"[:31]
def run_all(configurations, warmup_weeks, run_weeks, R, excel_path, input_dir):
    #Determine if excel path exists to know if we need to edit the new path or ratherr create a completely new file
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        # remove the default empty sheet openpyxl creates
        wb.remove(wb.active)

    total_weeks = warmup_weeks + run_weeks

    for config in configurations:
        n_slots, strategy, rule = config
        filename = os.path.join(input_dir, f"input-S{strategy}-{n_slots}.txt")
        sname    = sheet_name_for(config)

        print(f"\n{'='*60}")
        print(f"Config: {n_slots} urgent slots | Strategy {strategy} | Rule {rule}")
        print(f"  File : {filename}")
        print(f"  Sheet: {sname}")
        print(f"{'='*60}")

        if sname in wb.sheetnames: #if the sheet exists use it and just change after first row
            ws = wb[sname]
            print("  Sheet exists — appending rows.")

            for row in ws.iter_rows(min_row=2, max_col=6): #we delete (from row 2) what was written in the first 6 columns only
                for cell in row:
                    cell.value = None
        else: #if the sheet doesnt exist just make a new excel file/sheet
            ws = wb.create_sheet(sname)
            ws.append([ # columns we will apply
                "Replication",
                "Avg ElAppWT (h)",
                "Avg UrScanWT (h)",
                "Avg ElScanWT (h)",
                "Avg OT (h)",
                "Weighted Obj",
            ])

        # apply our simulation
        sim = Simulation(filename, total_weeks, R, rule)
        sim.setWeekSchedule()

        #applying multiple replicaitons
        for r in range(R):
            sim.resetSystem()
            random.seed(r)
            sim.runOneSimulation()

            # because we don't want to take into account a the warmup, slice the list
            post_el_app  = sim.movingAvgElectiveAppWT[warmup_weeks : warmup_weeks + run_weeks]
            post_ur_scan = sim.movingAvgUrgentScanWT [warmup_weeks : warmup_weeks + run_weeks]
            post_el_scan = sim.movingAvgElectiveScanWT[warmup_weeks: warmup_weeks + run_weeks]
            post_ot      = sim.movingAvgOT            [warmup_weeks : warmup_weeks + run_weeks]

            # now we can safely calculate the statistics (calculating average of replication via helperfunction), each row is thus the average over the replication (without warmup)
            avg_el_app  = safe_avg(post_el_app)
            avg_ur_scan = safe_avg(post_ur_scan)
            avg_el_scan = safe_avg(post_el_scan)
            avg_ot      = safe_avg(post_ot)
            weighted    = avg_el_app * sim.weightEl + avg_ur_scan * sim.weightUr

            # Write to excel
            row_idx = 2 + r  #always start at the second row
            ws.cell(row=row_idx, column=1, value=r)
            ws.cell(row=row_idx, column=2, value=round(avg_el_app,  5))
            ws.cell(row=row_idx, column=3, value=round(avg_ur_scan, 5))
            ws.cell(row=row_idx, column=4, value=round(avg_el_scan, 5))
            ws.cell(row=row_idx, column=5, value=round(avg_ot,      5))
            ws.cell(row=row_idx, column=6, value=round(weighted,    6))

            print(f"  r={r:>3}  elApp={avg_el_app:.3f}  urScan={avg_ur_scan:.3f}"
                  f"  elScan={avg_el_scan:.3f}  OT={avg_ot:.3f}  OV={weighted:.4f}")

        wb.save(excel_path)
        print(f"{excel_path}")

    print(f"\ndone")
if __name__ == "__main__":
    run_all(
        configurations=CONFIGURATIONS,
        warmup_weeks=WARMUP_WEEKS,
        run_weeks=RUN_WEEKS,
        R=R,
        excel_path=EXCEL_PATH,
        input_dir=INPUT_DIR,
    )
    