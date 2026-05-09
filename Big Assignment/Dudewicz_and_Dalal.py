import math
import random
import numpy as np
import os
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from simulation import Simulation

# ============================================================
# DUDEWICZ AND DALAL TWO-STAGE PROCEDURE — BATCH MEANS VERSION
# ============================================================

DESIGNS = [
    (12, 2, 4),
    (13, 2, 4),
    (12, 3, 4),
]

INPUT_DIR    = "Big Assignment/Inputs"
OUTPUT_EXCEL = "Big Assignment/Excel Files/Dudewicz_Dalal_results.xlsx"

N0           = 20      # First-stage batch count
P_STAR       = 0.90    # Desired probability of correct selection
SEED         = 1       # Single trajectory per design
WARMUP_WEEKS = 50
PILOT_WEEKS  = 2000

# d* floor: if the observed stage-1 gap is smaller than this,
# D_STAR_MIN is used instead to prevent N_i from exploding.
# Set to the smallest practically meaningful difference in your
# objective value (e.g. 1% of a typical mean).
D_STAR_MIN = 0.005



# ============================================================
# H-VALUE LOOKUP TABLE  (Dudewicz & Dalal 1975)
# ============================================================

H_TABLE = {
    (0.90, 20): {2: 1.896, 3: 2.342, 4: 2.583, 5: 2.747,
                 6: 2.870, 7: 2.969, 8: 3.051, 9: 3.121, 10: 3.182},
    (0.90, 40): {2: 1.852, 3: 2.283, 4: 2.514, 5: 2.669,
                 6: 2.785, 7: 2.878, 8: 2.954, 9: 3.019, 10: 3.076},
    (0.95, 20): {2: 2.453, 3: 2.872, 4: 3.101, 5: 3.258,
                 6: 3.377, 7: 3.472, 8: 3.551, 9: 3.619, 10: 3.679},
    (0.95, 40): {2: 2.386, 3: 2.786, 4: 3.003, 5: 3.150,
                 6: 3.260, 7: 3.349, 8: 3.422, 9: 3.484, 10: 3.539},
}


def lookup_h(p_star, n0, k):
    available_p = sorted(set(p for p, _ in H_TABLE))
    if p_star not in available_p:
        raise ValueError(f"P*={p_star} not in table. Available: {available_p}")

    available_n0 = sorted(n for p, n in H_TABLE if p == p_star)
    closest_n0   = min(available_n0, key=lambda x: abs(x - n0))
    if closest_n0 != n0:
        print(f"[h lookup] n0={n0} not in table for P*={p_star}. "
              f"Using closest: n0={closest_n0}.")

    row      = H_TABLE[(p_star, closest_n0)]
    k_capped = max(min(k, max(row)), min(row))
    if k_capped != k:
        print(f"[h lookup] k={k} outside table range. Using k={k_capped}.")

    return row[k_capped]


# ============================================================
# EXCEL STYLING
# ============================================================

BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
LIGHT_FILL  = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
GREEN2_FILL = PatternFill("solid", fgColor="70AD47")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FCE4D6")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")

WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT  = Font(name="Arial", bold=True, size=10)
REG_FONT   = Font(name="Arial", size=10)

thin        = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")


def style_header(cell, value, fill=BLUE_FILL):
    cell.value     = value
    cell.font      = WHITE_FONT
    cell.fill      = fill
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def style_value(cell, value, number_format=None, fill=None, bold=False, align=RIGHT):
    cell.value     = value
    cell.font      = BOLD_FONT if bold else REG_FONT
    cell.alignment = align
    cell.border    = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill


# ============================================================
# AUTOCORRELATION-BASED BATCH LENGTH
# ============================================================

def find_autocorr_lag(series):
    arr = np.array([v for v in series if math.isfinite(v)], dtype=float)
    if len(arr) < 10:
        return 1
    arr -= arr.mean()
    var = np.var(arr)
    if var == 0:
        return 1
    max_lag = min(len(arr) // 2, 200)
    for lag in range(1, max_lag + 1):
        ac = np.mean(arr[:-lag] * arr[lag:]) / var
        if abs(ac) < 0.01:
            return lag
    return max_lag


# ============================================================
# SIMULATION HELPERS
# ============================================================

def run_sim_get_series(input_file, rule, total_weeks, seed):
    sim = Simulation(input_file, total_weeks, 1, rule)
    sim.setWeekSchedule()
    sim.resetSystem()
    random.seed(seed)
    np.random.seed(seed)
    sim.runOneSimulation()

    series = []
    for w in range(WARMUP_WEEKS, len(sim.movingAvgElectiveAppWT)):
        ov = (
            sim.weightEl * sim.movingAvgElectiveAppWT[w]
            + sim.weightUr * sim.movingAvgUrgentScanWT[w]
        )
        if math.isfinite(ov):
            series.append(float(ov))
    return series


def determine_batch_size(input_file, rule):
    total_pilot = WARMUP_WEEKS + PILOT_WEEKS
    series      = run_sim_get_series(input_file, rule, total_pilot, seed=0)
    lag_ac      = find_autocorr_lag(series)
    M           = 5 * lag_ac
    return M, lag_ac


def extract_batch_means(series, M, num_batches):
    batch_means = []
    for b in range(num_batches):
        start = b * M
        end   = start + M
        if end > len(series):
            break
        batch_means.append(float(np.mean(series[start:end])))
    return batch_means


# ============================================================
# FIRST STAGE
# ============================================================

def first_stage(designs, n0):
    results = {}

    for idx, design in enumerate(designs):
        urgent, strategy, rule = design
        input_file = f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt"

        print("=" * 60)
        print(f"FIRST STAGE — SYSTEM {idx + 1}")
        print(f"  Design : urgent={urgent}, strategy={strategy}, rule={rule}")

        M, lag_ac = determine_batch_size(input_file, rule)
        print(f"  Lag    : {lag_ac}   Batch size M = {M}")

        total_weeks = WARMUP_WEEKS + n0 * M
        series      = run_sim_get_series(input_file, rule, total_weeks, seed=SEED)
        batch_means = extract_batch_means(series, M, n0)

        if len(batch_means) < n0:
            print(f"  WARNING: only {len(batch_means)} batches obtained (expected {n0}).")

        mean_s1 = float(np.mean(batch_means))
        var_s1  = float(np.var(batch_means, ddof=1))

        print(f"  Batches : {len(batch_means)}  mean = {mean_s1:.6f}  var = {var_s1:.8f}")

        results[idx] = {
            "design"            : design,
            "input_file"        : input_file,
            "batch_means_stage1": batch_means,
            "mean_stage1"       : mean_s1,
            "var_stage1"        : var_s1,
            "M"                 : M,
            "lag_ac"            : lag_ac,
        }

    return results


# ============================================================
# AUTO-DETERMINE d*  AND  LOOK UP h
# ============================================================

def determine_d_star_and_h(results, p_star, n0):
    """
    d* = gap between the two best stage-1 means, floored at D_STAR_MIN
         so that N_i never explodes when systems are very close.
    h  = D&D table lookup for (P*, n0, k).
    """
    k      = len(results)
    ranked = sorted(results.items(), key=lambda x: x[1]["mean_stage1"])

    best_mean        = ranked[0][1]["mean_stage1"]
    second_best_mean = ranked[1][1]["mean_stage1"]
    observed_gap     = abs(second_best_mean - best_mean)

    # Apply floor to prevent N_i explosion
    floor_active = observed_gap < D_STAR_MIN
    d_star       = max(observed_gap, D_STAR_MIN)

    h_value = lookup_h(p_star, n0, k)

    print("\n" + "=" * 60)
    print("INDIFFERENCE ZONE  &  H-VALUE")
    print(f"  Best system    : System {ranked[0][0]+1}  (mean = {best_mean:.6f})")
    print(f"  2nd best       : System {ranked[1][0]+1}  (mean = {second_best_mean:.6f})")
    print(f"  Observed gap   : {observed_gap:.8f}")
    print(f"  D_STAR_MIN     : {D_STAR_MIN:.8f}")
    if floor_active:
        print(f"  d* = D_STAR_MIN = {d_star:.8f}  (floor active — gap was smaller)")
    else:
        print(f"  d* = observed gap = {d_star:.8f}")
    print(f"  h  (P*={p_star}, n0={n0}, k={k}) : {h_value}")

    return d_star, h_value, observed_gap, floor_active


# ============================================================
# COMPUTE TOTAL SAMPLE SIZES
# ============================================================

def compute_total_sample_sizes(results, h_value, d_star, n0):
    print("\n" + "=" * 60)
    print("TOTAL SAMPLE SIZES")

    for idx in results:
        S2 = results[idx]["var_stage1"]
        Ni = max(
            n0 + 1,
            math.ceil((h_value ** 2 * S2) / (d_star ** 2)),
        )
        results[idx]["Ni"] = Ni
        print(f"  System {idx+1} -> N_i = {Ni}  (extra batches = {Ni - n0})")

    return results


# ============================================================
# SECOND STAGE
# ============================================================

def second_stage(results, n0, h_value, d_star):
    for idx in results:
        rule    = results[idx]["design"][2]
        M       = results[idx]["M"]
        Ni      = results[idx]["Ni"]
        n_extra = Ni - n0

        print("-" * 60)
        print(f"SECOND STAGE — SYSTEM {idx+1}  |  extra batches = {n_extra}")

        total_weeks        = WARMUP_WEEKS + Ni * M
        series             = run_sim_get_series(
            results[idx]["input_file"], rule, total_weeks, seed=SEED
        )
        all_batch_means    = extract_batch_means(series, M, Ni)
        stage2_batch_means = all_batch_means[n0:]

        stage2_mean = (
            float(np.mean(stage2_batch_means))
            if stage2_batch_means
            else results[idx]["mean_stage1"]
        )

        S2 = results[idx]["var_stage1"]
        if S2 <= 0:
            W1 = n0 / Ni
        else:
            inner = 1.0 - (Ni / n0) * (
                1.0 - (n_extra * d_star ** 2) / (h_value ** 2 * S2)
            )
            W1 = (n0 / Ni) * (1.0 + math.sqrt(max(inner, 0.0)))

        W2            = 1.0 - W1
        weighted_mean = W1 * results[idx]["mean_stage1"] + W2 * stage2_mean

        results[idx]["stage2_batch_means"] = stage2_batch_means
        results[idx]["all_batch_means"]    = all_batch_means
        results[idx]["stage2_mean"]        = stage2_mean
        results[idx]["W1"]                 = W1
        results[idx]["W2"]                 = W2
        results[idx]["weighted_mean"]      = weighted_mean

        print(f"  Stage-2 mean = {stage2_mean:.6f} | "
              f"W1 = {W1:.4f}  W2 = {W2:.4f} | "
              f"Weighted mean = {weighted_mean:.6f}")

    return results


# ============================================================
# BONFERRONI SIGNIFICANCE TESTS
# ============================================================
#
# (k-1) one-sided Welch t-tests on the full batch means:
#   H0: μ_best >= μ_j   H1: μ_best < μ_j
#
#   α_adj = (1 − P*) / (k − 1)   [Bonferroni correction]
#
# Best is SIGNIFICANT only if it beats every other system at α_adj.
# ============================================================

def bonferroni_significance(results, p_star):
    k         = len(results)
    alpha     = 1.0 - p_star
    alpha_adj = alpha / (k - 1)

    ranked   = sorted(results.items(), key=lambda x: x[1]["weighted_mean"])
    best_idx = ranked[0][0]
    best_bm  = np.array(results[best_idx]["all_batch_means"])

    print("\n" + "=" * 60)
    print("BONFERRONI SIGNIFICANCE  (one-sided Welch t-test)")
    print(f"  α = 1 − P* = {alpha:.2f}   k−1 = {k-1}   α_adj = {alpha_adj:.6f}")
    print(f"  Best system : System {best_idx + 1}")

    pairwise = {}
    all_sig  = True

    for idx, r in results.items():
        if idx == best_idx:
            continue

        other_bm      = np.array(r["all_batch_means"])
        t_stat, p_two = stats.ttest_ind(best_bm, other_bm, equal_var=False)
        p_one         = p_two / 2.0

        sig = bool(t_stat < 0 and p_one < alpha_adj)

        pairwise[idx] = {
            "t"        : float(t_stat),
            "p_one"    : float(p_one),
            "alpha_adj": alpha_adj,
            "sig"      : sig,
        }

        if not sig:
            all_sig = False

        marker = "significant" if sig else "NOT significant"
        print(f"  System {best_idx+1} vs System {idx+1}: "
              f"t = {t_stat:+.4f}  p = {p_one:.6f}  [{marker}]")

    verdict = "SIGNIFICANTLY BEST" if all_sig else "selected but NOT significant vs all"
    print(f"\n  Overall: System {best_idx+1} is {verdict} "
          f"at Bonferroni α_adj = {alpha_adj:.6f}")

    return best_idx, pairwise, all_sig, alpha_adj


# ============================================================
# SELECT BEST SYSTEM  (console)
# ============================================================

def select_best_system(results, all_sig):
    print("\n" + "=" * 60)
    print("FINAL RANKING  (minimise weighted mean)")
    print("=" * 60)

    ranked = sorted(results.items(), key=lambda x: x[1]["weighted_mean"])

    for rank, (idx, r) in enumerate(ranked, start=1):
        if rank == 1:
            sig_tag = " (SIGNIFICANT)" if all_sig else " (not significant)"
            print(f"  Rank {rank}  System {idx+1}  "
                  f"weighted mean = {r['weighted_mean']:.6f}  <- BEST{sig_tag}")
        else:
            print(f"  Rank {rank}  System {idx+1}  "
                  f"weighted mean = {r['weighted_mean']:.6f}")
        print(f"         Design = {r['design']}")


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_inputs_sheet(wb, results, d_star, h_value, observed_gap, floor_active):
    ws = wb.create_sheet("Inputs", 0)

    for col, w in enumerate([28, 28, 28, 28, 28, 28, 28, 30], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "Dudewicz & Dalal Two-Stage Procedure — Inputs (Batch Means)"
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill      = BLUE_FILL
    c.alignment = CENTER
    c.border    = THIN_BORDER

    k         = len(results)
    alpha_adj = (1 - P_STAR) / (k - 1)
    d_note    = f"{d_star:.8f}  {'← floor (D_STAR_MIN) active' if floor_active else '← observed gap'}"

    parameters = [
        ("n0  (first-stage batches)",        N0),
        ("P*  (target PCS)",                 P_STAR),
        ("D_STAR_MIN  (floor)",              D_STAR_MIN),
        ("Observed stage-1 gap",             observed_gap),
        ("d*  (used)",                       d_note),
        ("h   (from D&D table)",             h_value),
        ("k   (number of systems)",          k),
        ("Bonferroni α_adj = (1−P*)/(k−1)",  alpha_adj),
        ("Seed",                             SEED),
        ("Warmup weeks",                     WARMUP_WEEKS),
        ("Pilot weeks",                      PILOT_WEEKS),
        ("Input directory",                  INPUT_DIR),
    ]

    style_header(ws.cell(row=3, column=1), "Parameter")
    style_header(ws.cell(row=3, column=2), "Value")

    for row_i, (label, value) in enumerate(parameters, start=4):
        style_value(ws.cell(row=row_i, column=1), label, bold=True, align=LEFT)
        fmt = "#,##0.000000" if isinstance(value, float) else None
        fill = YELLOW_FILL if (label.startswith("d*") and floor_active) else None
        style_value(ws.cell(row=row_i, column=2), value, number_format=fmt, fill=fill)

    ds = 4 + len(parameters) + 2
    ws.merge_cells(start_row=ds, start_column=1, end_row=ds, end_column=8)
    style_value(ws.cell(row=ds, column=1),
                "Design Alternatives", fill=LIGHT_FILL, bold=True, align=CENTER)

    headers = ["System", "Urgent slots", "Strategy", "Rule",
               "Input file", "Lag L_ac", "Batch size M", "Stage-1 variance"]
    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=ds + 1, column=col), hdr)

    for idx, r in results.items():
        urgent, strategy, rule = r["design"]
        row = ds + 2 + idx
        vals = [idx + 1, urgent, strategy, rule,
                f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt",
                r["lag_ac"], r["M"], r["var_stage1"]]
        for col, v in enumerate(vals, start=1):
            style_value(
                ws.cell(row=row, column=col), v,
                number_format="#,##0.000000" if col == 8 else
                              "#,##0"        if col in [1, 2, 3, 4, 6, 7] else None,
                bold=(col == 1),
                align=CENTER if col in [1, 2, 3, 4, 6, 7] else
                      LEFT   if col == 5 else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_stage1_sheet(wb, results):
    ws = wb.create_sheet("Stage 1", 1)

    headers = ["System", "Batch #", "Batch mean X_ij",
               "Stage-1 mean", "Stage-1 variance", "Lag L_ac", "Batch size M"]
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1)
    t.value     = "First-Stage Batch Means"
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t.fill      = BLUE_FILL
    t.alignment = CENTER
    t.border    = THIN_BORDER

    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), hdr)

    row = 4
    for idx, r in results.items():
        for b_idx, bm in enumerate(r["batch_means_stage1"], start=1):
            zebra = GREY_FILL if row % 2 == 0 else None
            vals  = [idx + 1, b_idx, bm,
                     r["mean_stage1"], r["var_stage1"],
                     r["lag_ac"],      r["M"]]
            for col, v in enumerate(vals, start=1):
                style_value(
                    ws.cell(row=row, column=col), v,
                    number_format="#,##0"        if col in [1, 2, 6, 7] else
                                  "#,##0.000000" if col in [3, 4, 5] else None,
                    fill=zebra, bold=(col in [1, 2]),
                    align=CENTER if col in [1, 2, 6, 7] else RIGHT,
                )
            row += 1

    ws.freeze_panes = "A4"


def write_sample_size_sheet(wb, results, h_value, d_star, floor_active):
    ws = wb.create_sheet("Sample sizes", 2)

    headers = ["System", "S_i²(n0)", "h", "d* (used)",
               "Formula", "Calculated value", "N_i", "Extra batches N_i − n0"]
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 24

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1)
    t.value     = "Total Sample Size Calculation"
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t.fill      = BLUE_FILL
    t.alignment = CENTER
    t.border    = THIN_BORDER

    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), hdr)

    d_fill = YELLOW_FILL if floor_active else None

    for row_i, (idx, r) in enumerate(results.items(), start=4):
        S2   = r["var_stage1"]
        calc = (h_value ** 2 * S2) / (d_star ** 2)
        vals = [idx + 1, S2, h_value, d_star,
                "ceil(h² · S_i² / d*²)", calc, r["Ni"], r["Ni"] - N0]
        for col, v in enumerate(vals, start=1):
            style_value(
                ws.cell(row=row_i, column=col), v,
                number_format="#,##0"        if col in [1, 7, 8] else
                              "#,##0.000000" if col in [2, 3, 4, 6] else None,
                fill=ORANGE_FILL if col in [6, 7] else
                     d_fill      if col == 4 else None,
                bold=(col in [1, 7]),
                align=LEFT   if col == 5 else
                      CENTER if col in [1, 7, 8] else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_stage2_sheet(wb, results):
    ws = wb.create_sheet("Stage 2", 3)

    headers = ["System", "Batch #", "Batch mean X_ij",
               "Stage-2 mean", "W1", "W2", "Weighted mean X̄_i(N_i)"]
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1)
    t.value     = "Second-Stage Batch Means and Weighted Means"
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t.fill      = BLUE_FILL
    t.alignment = CENTER
    t.border    = THIN_BORDER

    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), hdr)

    row = 4
    for idx, r in results.items():
        bm_list = r.get("stage2_batch_means", [])
        if not bm_list:
            vals = [idx + 1, "—", "—",
                    r["stage2_mean"], r["W1"], r["W2"], r["weighted_mean"]]
            for col, v in enumerate(vals, start=1):
                style_value(
                    ws.cell(row=row, column=col), v,
                    number_format="#,##0.000000" if col in [4, 5, 6, 7] else None,
                    align=CENTER if col in [1, 2] else RIGHT,
                )
            row += 1
        else:
            for b_idx, bm in enumerate(bm_list, start=N0 + 1):
                zebra = GREY_FILL if row % 2 == 0 else None
                vals  = [idx + 1, b_idx, bm,
                         r["stage2_mean"], r["W1"], r["W2"], r["weighted_mean"]]
                for col, v in enumerate(vals, start=1):
                    style_value(
                        ws.cell(row=row, column=col), v,
                        number_format="#,##0"        if col in [1, 2] else
                                      "#,##0.000000",
                        fill=zebra, bold=(col in [1, 2]),
                        align=CENTER if col in [1, 2] else RIGHT,
                    )
                row += 1

    ws.freeze_panes = "A4"


def write_significance_sheet(wb, results, best_idx, pairwise, all_sig, alpha_adj):
    ws = wb.create_sheet("Significance", 4)

    for col, w in enumerate([22, 22, 18, 18, 20, 24, 26], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1)
    t.value     = "Bonferroni-Corrected Pairwise Significance Tests (one-sided Welch t-test)"
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t.fill      = BLUE_FILL
    t.alignment = CENTER
    t.border    = THIN_BORDER

    ws.merge_cells("A3:G3")
    note = ws.cell(row=3, column=1)
    note.value = (
        f"H0: best (System {best_idx+1}) is NOT better than system j   |   "
        f"H1: best < system j   |   "
        f"α_adj = (1−P*) / (k−1) = {alpha_adj:.6f}"
    )
    note.font      = Font(name="Arial", italic=True, size=9)
    note.alignment = LEFT
    note.border    = THIN_BORDER
    note.fill      = LIGHT_FILL

    headers = ["Best system", "vs. System", "N_best batches",
               "N_other batches", "t-statistic",
               "p-value (one-sided)", "Significant at α_adj?"]
    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=5, column=col), hdr)

    row = 6
    for j_idx, pw in pairwise.items():
        sig   = pw["sig"]
        label = "YES" if sig else "NO"
        color = GREEN2_FILL if sig else YELLOW_FILL
        vals  = [
            f"System {best_idx + 1}",
            f"System {j_idx + 1}",
            len(results[best_idx]["all_batch_means"]),
            len(results[j_idx]["all_batch_means"]),
            pw["t"],
            pw["p_one"],
            label,
        ]
        for col, v in enumerate(vals, start=1):
            style_value(
                ws.cell(row=row, column=col), v,
                number_format="#,##0"        if col in [3, 4] else
                              "#,##0.000000" if col in [5, 6] else None,
                fill=color if col == 7 else None,
                bold=(col == 7),
                align=CENTER,
            )
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    verdict = ws.cell(row=row, column=1)
    if all_sig:
        verdict.value = (
            f"SIGNIFICANTLY BEST — System {best_idx+1} beats all other systems "
            f"at Bonferroni-corrected α_adj = {alpha_adj:.6f}"
        )
        verdict.fill = GREEN2_FILL
        verdict.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    else:
        verdict.value = (
            f"NOT SIGNIFICANT — System {best_idx+1} is selected as best but does NOT "
            f"significantly beat all others at Bonferroni-corrected α_adj = {alpha_adj:.6f}"
        )
        verdict.fill = YELLOW_FILL
        verdict.font = Font(name="Arial", bold=True, size=11)
    verdict.alignment = CENTER
    verdict.border    = THIN_BORDER

    ws.freeze_panes = "A6"


def write_summary_sheet(wb, results, best_idx, pairwise, all_sig, alpha_adj):
    ws = wb.create_sheet("Final summary", 5)

    headers = ["System", "Design", "n0", "N_i",
               "Stage-1 mean", "Stage-2 mean",
               "W1", "W2", "Weighted mean", "Rank",
               "Best?", "Significant? (Bonferroni)"]
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1)
    t.value     = "Final Dudewicz & Dalal Selection Summary"
    t.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    t.fill      = BLUE_FILL
    t.alignment = CENTER
    t.border    = THIN_BORDER

    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), hdr)

    ranked   = sorted(results.items(), key=lambda x: x[1]["weighted_mean"])
    ranks    = {idx: rank for rank, (idx, _) in enumerate(ranked, start=1)}

    for row_i, (idx, r) in enumerate(results.items(), start=4):
        is_best = (idx == best_idx)

        if is_best:
            sig_label = "YES — significant" if all_sig else "YES — NOT significant"
        else:
            sig_label = "—"

        vals = [
            idx + 1,
            str(r["design"]),
            N0,
            r["Ni"],
            r["mean_stage1"],
            r["stage2_mean"],
            r["W1"],
            r["W2"],
            r["weighted_mean"],
            ranks[idx],
            "YES" if is_best else "NO",
            sig_label,
        ]

        for col, v in enumerate(vals, start=1):
            fill = None
            if col == 11:
                fill = GREEN_FILL if is_best else RED_FILL
            if col == 12 and is_best:
                fill = GREEN2_FILL if all_sig else YELLOW_FILL
            if col == 9:
                fill = ORANGE_FILL

            style_value(
                ws.cell(row=row_i, column=col), v,
                number_format="#,##0"        if col in [1, 3, 4, 10] else
                              "#,##0.000000" if col in [5, 6, 7, 8, 9] else None,
                fill=fill,
                bold=(col in [1, 9, 10, 11, 12]),
                align=LEFT   if col == 2 else
                      CENTER if col in [1, 3, 4, 10, 11, 12] else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_excel_results(results, d_star, h_value, observed_gap, floor_active,
                        best_idx, pairwise, all_sig, alpha_adj):
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    write_inputs_sheet(wb, results, d_star, h_value, observed_gap, floor_active)
    write_stage1_sheet(wb, results)
    write_sample_size_sheet(wb, results, h_value, d_star, floor_active)
    write_stage2_sheet(wb, results)
    write_significance_sheet(wb, results, best_idx, pairwise, all_sig, alpha_adj)
    write_summary_sheet(wb, results, best_idx, pairwise, all_sig, alpha_adj)

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved → {OUTPUT_EXCEL}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("\nDUDEWICZ & DALAL — BATCH MEANS VERSION")
    print(f"P* = {P_STAR}   n0 = {N0}   D_STAR_MIN = {D_STAR_MIN}   seed = {SEED}")
    print("=" * 60)

    # Stage 1
    results = first_stage(DESIGNS, N0)

    # Auto d* + h lookup
    d_star, h_value, observed_gap, floor_active = determine_d_star_and_h(
        results, P_STAR, N0
    )

    # N_i per system
    results = compute_total_sample_sizes(results, h_value, d_star, N0)

    # Stage 2
    results = second_stage(results, N0, h_value, d_star)

    # Bonferroni significance
    best_idx, pairwise, all_sig, alpha_adj = bonferroni_significance(results, P_STAR)

    # Console summary
    select_best_system(results, all_sig)

    # Excel
    write_excel_results(results, d_star, h_value, observed_gap, floor_active,
                        best_idx, pairwise, all_sig, alpha_adj)


if __name__ == "__main__":
    main()
