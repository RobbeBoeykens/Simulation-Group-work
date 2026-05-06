

import math
import random
import numpy as np

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from simulation import Simulation

# ============================================================
# DUDEWICZ AND DALAL TWO-STAGE PROCEDURE
# ============================================================
# User inputs
# ============================================================

DESIGNS = [
    # (urgent_slots, strategy, rule)
    (13, 2, 4),
    (12, 3, 3),
]

INPUT_DIR = "Big Assignment/Inputs"

OUTPUT_EXCEL = "Big Assignment/Excel Files/Dudewicz_Dalal_results.xlsx"

# First-stage sample size n0
N0 = 20

# Indifference-zone parameter d*
D_STAR = 0.005

# h-value from the Dudewicz & Dalal table
H_VALUE = 1.896

# Warmup period
WARMUP_WEEKS = 50

# Pilot length for autocorrelation estimation
PILOT_WEEKS = 2000

# ============================================================
# EXCEL STYLING
# ============================================================

BLUE_FILL = PatternFill("solid", fgColor="1F4E79")
LIGHT_FILL = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")

WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
REG_FONT = Font(name="Arial", size=10)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header(cell, value, fill=BLUE_FILL):
    cell.value = value
    cell.font = WHITE_FONT
    cell.fill = fill
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_value(cell, value, number_format=None, fill=None, bold=False, align=RIGHT):
    cell.value = value
    cell.font = BOLD_FONT if bold else REG_FONT
    cell.alignment = align
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill


# ============================================================
# AUTOCORRELATION-BASED BATCH LENGTH
# ============================================================

def find_autocorr_lag(series):
    """
    Find the first lag where autocorrelation becomes small.
    Batch size M = 5 * L_ac.
    """

    arr = np.array([v for v in series if math.isfinite(v)], dtype=float)

    if len(arr) < 10:
        return 1

    arr = arr - arr.mean()
    var = np.var(arr)

    if var == 0:
        return 1

    max_lag = min(len(arr) // 2, 200)

    for lag in range(1, max_lag + 1):
        ac = np.mean(arr[:-lag] * arr[lag:]) / var

        if abs(ac) < 0.01:
            return lag

    return max_lag


# ============================================================
# SIMULATION HELPERS
# ============================================================

def run_long_sim(input_file, rule, total_weeks, seed):
    sim = Simulation(input_file, total_weeks, 1, rule)
    sim.setWeekSchedule()
    sim.resetSystem()

    random.seed(seed)
    np.random.seed(seed)

    sim.runOneSimulation()

    return sim


def get_weekly_objective_series(sim):
    """
    Weekly objective value:
    OV = weightEl * elective waiting time
       + weightUr * urgent waiting time
    """

    series = []

    for w in range(WARMUP_WEEKS, len(sim.movingAvgElectiveAppWT)):
        ov = (
            sim.weightEl * sim.movingAvgElectiveAppWT[w]
            + sim.weightUr * sim.movingAvgUrgentScanWT[w]
        )

        if math.isfinite(ov):
            series.append(float(ov))

    return series


# ============================================================
# BATCH MEANS
# ============================================================

def compute_batch_mean(series, M):
    """
    Compute one replication output using batch means.
    The simulation output is divided into batches of size M.
    The replication result is the average of all batch means.
    """

    num_batches = len(series) // M

    if num_batches == 0:
        return float(np.mean(series))

    batch_means = []

    for b in range(num_batches):
        start = b * M
        end = (b + 1) * M

        batch = series[start:end]
        batch_means.append(float(np.mean(batch)))

    return float(np.mean(batch_means))


# ============================================================
# DETERMINE BATCH SIZE USING AUTOCORRELATION
# ============================================================

def determine_batch_size(input_file, rule):

    total_pilot_weeks = WARMUP_WEEKS + PILOT_WEEKS

    sim = run_long_sim(
        input_file=input_file,
        rule=rule,
        total_weeks=total_pilot_weeks,
        seed=0,
    )

    pilot_series = get_weekly_objective_series(sim)

    lag_ac = find_autocorr_lag(pilot_series)

    M = 5 * lag_ac

    return M, lag_ac


# ============================================================
# ONE REPLICATION
# ============================================================

def simulate_replication(urgent_slots, strategy, rule, replication_id):

    input_file = f"{INPUT_DIR}/input-S{strategy}-{urgent_slots}.txt"

    M, lag_ac = determine_batch_size(input_file, rule)

    total_weeks = WARMUP_WEEKS + 10 * M

    sim = run_long_sim(
        input_file=input_file,
        rule=rule,
        total_weeks=total_weeks,
        seed=replication_id,
    )

    series = get_weekly_objective_series(sim)

    replication_value = compute_batch_mean(series, M)

    return replication_value, M, lag_ac


# ============================================================
# FIRST STAGE
# ============================================================

def first_stage(designs, n0):

    results = {}

    for idx, design in enumerate(designs):

        urgent, strategy, rule = design

        print("=" * 60)
        print(f"SYSTEM {idx + 1}")
        print(f"Urgent slots = {urgent}")
        print(f"Strategy     = {strategy}")
        print(f"Rule         = {rule}")

        replication_outputs = []

        for r in range(n0):

            value, M, lag_ac = simulate_replication(
                urgent_slots=urgent,
                strategy=strategy,
                rule=rule,
                replication_id=r,
            )

            replication_outputs.append(value)

            print(
                f"Replication {r + 1:2d} | "
                f"Value = {value:.6f} | "
                f"Lag = {lag_ac} | "
                f"M = {M}"
            )

        mean_value = float(np.mean(replication_outputs))
        variance_value = float(np.var(replication_outputs, ddof=1))

        results[idx] = {
            "design": design,
            "outputs": replication_outputs,
            "mean_stage1": mean_value,
            "var_stage1": variance_value,
            "M": M,
            "lag_ac": lag_ac,
        }

        print(f"Stage-1 mean     = {mean_value:.6f}")
        print(f"Stage-1 variance = {variance_value:.8f}")

    return results


# ============================================================
# COMPUTE TOTAL SAMPLE SIZE N_i
# ============================================================

def compute_total_sample_sizes(results, h_value, d_star, n0):

    for idx in results:

        S2 = results[idx]["var_stage1"]

        Ni = max(
            n0 + 1,
            math.ceil((h_value ** 2 * S2) / (d_star ** 2))
        )

        results[idx]["Ni"] = Ni

        print(f"System {idx + 1} -> Total sample size Ni = {Ni}")

    return results


# ============================================================
# SECOND STAGE
# ============================================================

def second_stage(results, n0):

    for idx in results:

        urgent, strategy, rule = results[idx]["design"]

        current_outputs = results[idx]["outputs"]

        Ni = results[idx]["Ni"]

        additional_replications = Ni - n0

        second_stage_outputs = []

        print("-" * 60)
        print(f"SECOND STAGE SYSTEM {idx + 1}")
        print(f"Additional replications = {additional_replications}")

        for r in range(additional_replications):

            replication_id = n0 + r

            value, _, _ = simulate_replication(
                urgent_slots=urgent,
                strategy=strategy,
                rule=rule,
                replication_id=replication_id,
            )

            second_stage_outputs.append(value)

            print(
                f"Additional replication {replication_id + 1:2d} | "
                f"Value = {value:.6f}"
            )

        stage2_mean = (
            float(np.mean(second_stage_outputs))
            if len(second_stage_outputs) > 0
            else current_outputs[-1]
        )

        all_outputs = current_outputs + second_stage_outputs
        overall_mean = float(np.mean(all_outputs))
        overall_variance = float(np.var(all_outputs, ddof=1)) if len(all_outputs) > 1 else 0.0

        # Dudewicz & Dalal weights from the lecture slide:
        # W_i1 = (n0 / Ni) * (1 + sqrt(1 - (Ni / n0) *
        #        (1 - ((Ni - n0) * d_star^2) / (h^2 * S_i^2(n0)))))
        # W_i2 = 1 - W_i1
        S2 = results[idx]["var_stage1"]

        if S2 <= 0:
            W1 = n0 / Ni
        else:
            inside_sqrt = 1.0 - (Ni / n0) * (
                1.0 - ((Ni - n0) * (D_STAR ** 2)) / ((H_VALUE ** 2) * S2)
            )
            inside_sqrt = max(inside_sqrt, 0.0)
            W1 = (n0 / Ni) * (1.0 + math.sqrt(inside_sqrt))

        W2 = 1.0 - W1

        weighted_mean = (
            W1 * results[idx]["mean_stage1"]
            + W2 * stage2_mean
        )

        results[idx]["stage2_outputs"] = second_stage_outputs
        results[idx]["all_outputs"] = all_outputs
        results[idx]["stage2_mean"] = stage2_mean
        results[idx]["overall_mean"] = overall_mean
        results[idx]["overall_variance"] = overall_variance
        results[idx]["W1"] = W1
        results[idx]["W2"] = W2
        results[idx]["weighted_mean"] = weighted_mean

        print(f"Stage-2 mean  = {stage2_mean:.6f}")
        print(f"W1            = {W1:.6f}")
        print(f"W2            = {W2:.6f}")
        print(f"Weighted mean = {weighted_mean:.6f}")

    return results


# ============================================================
# SELECT BEST SYSTEM
# ============================================================

def select_best_system(results):

    best_idx = None
    best_value = float("inf")

    print("\n" + "=" * 60)
    print("FINAL RANKING")
    print("=" * 60)

    for idx in results:

        value = results[idx]["weighted_mean"]

        print(
            f"System {idx + 1} | "
            f"Weighted mean = {value:.6f}"
        )

        if value < best_value:
            best_value = value
            best_idx = idx

    print("\nBEST SYSTEM")
    print(f"System {best_idx + 1}")
    print(f"Design = {results[best_idx]['design']}")
    print(f"Objective value = {best_value:.6f}")


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_inputs_sheet(wb, results):
    ws = wb.create_sheet("Inputs", 0)

    for col, width in enumerate([18, 18, 18, 18, 18, 18, 18, 22], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Dudewicz & Dalal Two-Stage Procedure — Inputs"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    parameters = [
        ("n0", N0),
        ("d*", D_STAR),
        ("h", H_VALUE),
        ("Warmup weeks", WARMUP_WEEKS),
        ("Pilot weeks", PILOT_WEEKS),
        ("Input directory", INPUT_DIR),
    ]

    style_header(ws.cell(row=3, column=1), "Parameter")
    style_header(ws.cell(row=3, column=2), "Value")

    for row_idx, (label, value) in enumerate(parameters, start=4):
        style_value(ws.cell(row=row_idx, column=1), label, bold=True, align=LEFT)
        style_value(ws.cell(row=row_idx, column=2), value, number_format="#,##0.00000" if isinstance(value, float) else None)

    design_start = 4 + len(parameters) + 2
    ws.merge_cells(start_row=design_start, start_column=1, end_row=design_start, end_column=8)
    style_value(ws.cell(row=design_start, column=1), "Design Alternatives", fill=LIGHT_FILL, bold=True, align=CENTER)

    headers = ["System", "Urgent slots", "Strategy", "Rule", "Input file", "Lag L_ac", "Batch length M", "First-stage variance"]
    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=design_start + 1, column=col), header)

    for idx, result in results.items():
        urgent, strategy, rule = result["design"]
        row = design_start + 2 + idx
        values = [
            idx + 1,
            urgent,
            strategy,
            rule,
            f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt",
            result["lag_ac"],
            result["M"],
            result["var_stage1"],
        ]
        for col, value in enumerate(values, start=1):
            style_value(
                ws.cell(row=row, column=col),
                value,
                number_format="#,##0.000000" if col == 8 else "#,##0" if col in [1, 2, 3, 4, 6, 7] else None,
                bold=col == 1,
                align=CENTER if col in [1, 2, 3, 4, 6, 7] else LEFT if col == 5 else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_stage1_sheet(wb, results):
    ws = wb.create_sheet("Stage 1", 1)

    headers = ["System", "Replication", "Output X_ij", "Stage-1 mean", "Stage-1 variance", "Lag L_ac", "Batch length M"]
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "First-Stage Replication Outputs and Calculations"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    row = 4
    for idx, result in results.items():
        for rep_idx, value in enumerate(result["outputs"], start=1):
            zebra = GREY_FILL if row % 2 == 0 else None
            values = [
                idx + 1,
                rep_idx,
                value,
                result["mean_stage1"],
                result["var_stage1"],
                result["lag_ac"],
                result["M"],
            ]
            for col, val in enumerate(values, start=1):
                style_value(
                    ws.cell(row=row, column=col),
                    val,
                    number_format="#,##0" if col in [1, 2, 6, 7] else "#,##0.000000",
                    fill=zebra,
                    bold=col in [1, 2],
                    align=CENTER if col in [1, 2, 6, 7] else RIGHT,
                )
            row += 1

    ws.freeze_panes = "A4"


def write_sample_size_sheet(wb, results):
    ws = wb.create_sheet("Sample sizes", 2)

    headers = ["System", "S_i^2(n0)", "h", "d*", "Formula", "Calculated value", "N_i", "Additional reps N_i - n0"]
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 24

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "Total Sample Size Calculation"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    for row_idx, (idx, result) in enumerate(results.items(), start=4):
        S2 = result["var_stage1"]
        calculated_value = (H_VALUE ** 2 * S2) / (D_STAR ** 2) if D_STAR != 0 else 0.0
        values = [
            idx + 1,
            S2,
            H_VALUE,
            D_STAR,
            "ceil(h^2 * S_i^2 / d*^2)",
            calculated_value,
            result["Ni"],
            result["Ni"] - N0,
        ]
        for col, val in enumerate(values, start=1):
            style_value(
                ws.cell(row=row_idx, column=col),
                val,
                number_format="#,##0" if col in [1, 7, 8] else "#,##0.000000",
                fill=ORANGE_FILL if col in [6, 7] else None,
                bold=col in [1, 7],
                align=LEFT if col == 5 else CENTER if col in [1, 7, 8] else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_stage2_sheet(wb, results):
    ws = wb.create_sheet("Stage 2", 3)

    headers = ["System", "Additional replication", "Output X_ij", "Stage-2 mean", "W1", "W2", "Weighted mean"]
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "Second-Stage Replication Outputs and Weighted Means"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    row = 4
    for idx, result in results.items():
        if len(result["stage2_outputs"]) == 0:
            values = [idx + 1, "No extra reps", "", result["stage2_mean"], result["W1"], result["W2"], result["weighted_mean"]]
            for col, val in enumerate(values, start=1):
                style_value(
                    ws.cell(row=row, column=col),
                    val,
                    number_format="#,##0.000000" if col in [4, 5, 6, 7] else None,
                    align=CENTER if col in [1, 2] else RIGHT,
                )
            row += 1
        else:
            for rep_idx, value in enumerate(result["stage2_outputs"], start=N0 + 1):
                zebra = GREY_FILL if row % 2 == 0 else None
                values = [idx + 1, rep_idx, value, result["stage2_mean"], result["W1"], result["W2"], result["weighted_mean"]]
                for col, val in enumerate(values, start=1):
                    style_value(
                        ws.cell(row=row, column=col),
                        val,
                        number_format="#,##0" if col in [1, 2] else "#,##0.000000",
                        fill=zebra,
                        bold=col in [1, 2],
                        align=CENTER if col in [1, 2] else RIGHT,
                    )
                row += 1

    ws.freeze_panes = "A4"


def write_summary_sheet(wb, results):
    ws = wb.create_sheet("Final summary", 4)

    headers = ["System", "Design", "n0", "N_i", "Stage-1 mean", "Stage-2 mean", "W1", "W2", "Weighted mean", "Rank", "Selected best?"]
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "Final Dudewicz & Dalal Selection Summary"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    sorted_items = sorted(results.items(), key=lambda item: item[1]["weighted_mean"])
    ranks = {idx: rank for rank, (idx, _) in enumerate(sorted_items, start=1)}
    best_idx = sorted_items[0][0]

    for row_idx, (idx, result) in enumerate(results.items(), start=4):
        values = [
            idx + 1,
            str(result["design"]),
            N0,
            result["Ni"],
            result["mean_stage1"],
            result["stage2_mean"],
            result["W1"],
            result["W2"],
            result["weighted_mean"],
            ranks[idx],
            "YES" if idx == best_idx else "NO",
        ]
        for col, val in enumerate(values, start=1):
            fill = None
            if col == 11:
                fill = GREEN_FILL if val == "YES" else RED_FILL
            if col == 9:
                fill = ORANGE_FILL
            style_value(
                ws.cell(row=row_idx, column=col),
                val,
                number_format="#,##0" if col in [1, 3, 4, 10] else "#,##0.000000" if col in [5, 6, 7, 8, 9] else None,
                fill=fill,
                bold=col in [1, 9, 10, 11],
                align=LEFT if col == 2 else CENTER if col in [1, 3, 4, 10, 11] else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_excel_results(results):
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    write_inputs_sheet(wb, results)
    write_stage1_sheet(wb, results)
    write_sample_size_sheet(wb, results)
    write_stage2_sheet(wb, results)
    write_summary_sheet(wb, results)

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved -> {OUTPUT_EXCEL}")


# ============================================================
# MAIN
# ============================================================

def main():

    print("\nDUDEWICZ & DALAL TWO-STAGE PROCEDURE")
    print("=" * 60)

    # --------------------------------------------------------
    # FIRST STAGE
    # --------------------------------------------------------

    results = first_stage(
        designs=DESIGNS,
        n0=N0,
    )

    # --------------------------------------------------------
    # COMPUTE TOTAL SAMPLE SIZES
    # --------------------------------------------------------

    results = compute_total_sample_sizes(
        results=results,
        h_value=H_VALUE,
        d_star=D_STAR,
        n0=N0,
    )

    # --------------------------------------------------------
    # SECOND STAGE
    # --------------------------------------------------------

    results = second_stage(
        results=results,
        n0=N0,
    )

    # --------------------------------------------------------
    # SELECT BEST SYSTEM
    # --------------------------------------------------------

    select_best_system(results)

    # --------------------------------------------------------
    # WRITE EXCEL OUTPUT
    # --------------------------------------------------------

    write_excel_results(results)


if __name__ == "__main__":
    main()