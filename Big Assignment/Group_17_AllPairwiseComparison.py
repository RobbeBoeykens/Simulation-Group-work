import os
import math
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from simulation import Simulation

# ============================================================
# SETTINGS — enter the four designs here
# ============================================================
# Each design is one alternative a.
# The input file is chosen as: input-S{strategy}-{urgent_slots}.txt
# The appointment rule is passed separately to Simulation(..., rule).

OUTPUT_EXCEL = "Big Assignment/Excel Files/All_pairwise_comparison.xlsx"
INPUT_DIR = "Big Assignment/Inputs"

DESIGNS = [
    {"name": "a1", "urgent_slots": 12, "strategy": 2, "rule": 4},
    {"name": "a2", "urgent_slots": 13, "strategy": 2, "rule": 4},
    {"name": "a3", "urgent_slots": 14, "strategy": 2, "rule": 4},
    {"name": "a4", "urgent_slots": 12, "strategy": 3, "rule": 4},
    {"name": "a5", "urgent_slots": 13, "strategy": 3, "rule": 4},
    {"name": "a6", "urgent_slots": 14, "strategy": 3, "rule": 4},
]

#
# One long steady-state trajectory per design.
# The batches themselves become the observations.
N_REPLICATIONS = 1

# Simulation length.
WARMUP_WEEKS = 50

# Batch mean settings.
# First estimate the autocorrelation lag L_ac with a pilot run.
# Then set the batch length M = 5 * L_ac, exactly like in the batch mean file.
BATCHES_PER_REPLICATION = 8
PILOT_WEEKS = 2000
FORCE_M = None  # None = automatic M from autocorrelation; int = fixed M

# One common seed for the single long trajectory.
# All designs use the same random stream (CRN).
BASE_SEED = 10_000

# Overall confidence level voor de family of pairwise comparisons.
# Met k = 4 designs zijn er c = k(k-1)/2 = 6 paren.
# Bonferroni: elk individueel interval op niveau 1 - alpha/c.
OVERALL_CONFIDENCE = 0.95

# ============================================================
# Excel styling
# ============================================================
BLUE_FILL = PatternFill("solid", fgColor="1F4E79")
LIGHT_FILL = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
GREY_FILL = PatternFill("solid", fgColor="F2F2F2")

WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
REG_FONT = Font(name="Arial", size=10)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


# ============================================================
# Helper functions
# ============================================================
def safe_mean(values):
    """Mean over finite values. Logs hoeveel observaties geweerd werden."""
    valid = [float(v) for v in values if math.isfinite(float(v))]
    if len(valid) < len(values):
        dropped = len(values) - len(valid)
        print(f"  [warn] {dropped} non-finite value(s) dropped from mean")
    return sum(valid) / len(valid) if valid else 0.0


def safe_std(values):
    valid = [float(v) for v in values if math.isfinite(float(v))]
    if len(valid) < 2:
        return 0.0
    return float(np.std(np.array(valid, dtype=float), ddof=1))


# 95% en 99.17% (= 1 - 0.05/6) kritieke t-waarden voor Bonferroni met c = 6 paren.
# Voor een algemene confidence van 1 - alpha = 0.95 over 6 intervals
# moet elk interval op niveau 1 - alpha/c geconstrueerd worden.
T_CRIT_TABLE = {
    0.95: {  # tweezijdige 95%
        1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571, 6: 2.447,
        7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228, 11: 2.201, 12: 2.179,
        13: 2.160, 14: 2.145, 15: 2.131, 16: 2.120, 17: 2.110, 18: 2.101,
        19: 2.093, 20: 2.086, 21: 2.080, 22: 2.074, 23: 2.069, 24: 2.064,
        25: 2.060, 26: 2.056, 27: 2.052, 28: 2.048, 29: 2.045, 30: 2.042,
        40: 2.021, 60: 2.000, 120: 1.980,
    },
    # Tweezijdige (1 - 0.05/6) = 99.1667%-CI, alpha/2 = 0.004167 in elke staart.
    # Waarden uit standaard t-tabellen (kolom 0.005 als conservatieve benadering;
    # exacte waarden via scipy.stats.t.ppf(1 - 0.05/12, df) zouden iets kleiner zijn).
    0.991667: {
        1: 76.39, 2: 9.925, 3: 5.841, 4: 4.604, 5: 4.032, 6: 3.707,
        7: 3.499, 8: 3.355, 9: 3.250, 10: 3.169, 11: 3.106, 12: 3.055,
        13: 3.012, 14: 2.977, 15: 2.947, 16: 2.921, 17: 2.898, 18: 2.878,
        19: 2.861, 20: 2.845, 21: 2.831, 22: 2.819, 23: 2.807, 24: 2.797,
        25: 2.787, 26: 2.779, 27: 2.771, 28: 2.763, 29: 2.756, 30: 2.750,
        40: 2.704, 60: 2.660, 120: 2.617,
    },
}


def t_critical(df, confidence):
    """Geeft tweezijdige t-kritieke waarde voor gegeven df en confidence niveau."""
    table = T_CRIT_TABLE.get(confidence)
    if table is None:
        return 1.96  # fallback normaalbenadering
    if df in table:
        return table[df]
    # Pak dichtstbijzijnde lagere df, dit is conservatiever omdat de t-waarde hoger is.
    available = sorted(table.keys())
    for d in reversed(available):
        if d <= df:
            return table[d]
    return table[available[0]]


def ci_half(values, confidence):
    """CI half-width via paired-t (of one-sample-t) op gewenst confidence niveau."""
    valid = [float(v) for v in values if math.isfinite(float(v))]
    n = len(valid)
    if n < 2:
        return 0.0
    df = n - 1
    t_crit = t_critical(df, confidence)
    return float(t_crit * safe_std(valid) / math.sqrt(n))


def find_autocorr_lag(series):
    """
    Zoek de kleinste lag L_ac waarvoor |autocorr(lag)| < 0.01.
    Slide 2: batchlengte M = 5 * L_ac.
    Geeft minimaal 1 terug.
    """
    arr = np.array([v for v in series if math.isfinite(v)], dtype=float)
    if len(arr) < 10:
        return 1

    arr = arr - arr.mean()
    var = np.var(arr)
    if var == 0:
        return 1

    max_lag = min(len(arr) // 2, 200)
    for lag in range(1, max_lag + 1):
        ac = np.mean(arr[:-lag] * arr[lag:]) / var
        if abs(ac) < 0.01:
            return lag
    return max_lag


def compute_batch_means(series, M, L):
    """
    Verdeelt `series` in L batches van M weken.
    Geeft lijst van L batchgemiddelden terug.
    """
    batch_means = []
    for l in range(L):
        batch = series[l * M : (l + 1) * M]
        batch_means.append(safe_mean(batch))
    return batch_means


# === Variance reduction helpers ===

def safe_np_var(values):
    arr = np.array(values, dtype=float)
    if len(arr) < 2:
        return 0.0
    return float(np.var(arr, ddof=1))


def var_reduction_pct(new_values, raw_values):
    raw_var = safe_np_var(raw_values)
    new_var = safe_np_var(new_values)
    if raw_var <= 0:
        return 0.0
    return 100.0 * (1.0 - new_var / raw_var)


def get_input_file(design):
    return f"{INPUT_DIR}/input-S{design['strategy']}-{design['urgent_slots']}.txt"


def run_simulation(design, seed, total_weeks, use_antithetic=False):
    """
    Runs one simulation. If use_antithetic=True, random.random and random.randint
    are mirrored to create antithetic variates, as in the screening script.
    """
    input_file = get_input_file(design)

    sim = Simulation(input_file, total_weeks, 1, design["rule"])
    sim.setWeekSchedule()
    sim.resetSystem()
    random.seed(seed)

    if use_antithetic:
        original_random = random.random
        original_randint = random.randint

        random.random = lambda: 1.0 - original_random()

        def anti_randint(a, b):
            k = original_randint(a, b)
            return a + b - k

        random.randint = anti_randint

        try:
            sim.runOneSimulation()
        finally:
            random.random = original_random
            random.randint = original_randint
    else:
        sim.runOneSimulation()

    return sim


def get_weekly_objective_values(sim, total_weeks):
    """
    Returns post-warmup weekly objective values.
    """
    weekly_objective_values = []
    for week in range(WARMUP_WEEKS, total_weeks):
        elective_app_wt = sim.movingAvgElectiveAppWT[week]
        urgent_scan_wt = sim.movingAvgUrgentScanWT[week]

        if not math.isfinite(elective_app_wt):
            elective_app_wt = 0.0
        if not math.isfinite(urgent_scan_wt):
            urgent_scan_wt = 0.0

        objective_value = sim.weightEl * elective_app_wt + sim.weightUr * urgent_scan_wt
        weekly_objective_values.append(objective_value)

    return weekly_objective_values


def get_post_warmup_metric_series(sim, attr_name, total_weeks):
    values = getattr(sim, attr_name)
    out = []
    for week in range(WARMUP_WEEKS, total_weeks):
        value = values[week]
        out.append(value if math.isfinite(value) else 0.0)
    return out


def count_scanned_by_batch(sim, M, L, patient_type):
    """
    Control variate Y per batch: number of scanned patients of type patient_type
    with scanWeek inside the batch after warmup.
    patient_type 1 = elective, patient_type 2 = urgent.
    """
    counts = []
    for batch_idx in range(L):
        start_week = WARMUP_WEEKS + batch_idx * M
        end_week = WARMUP_WEEKS + (batch_idx + 1) * M
        count = sum(
            1 for patient in sim.patients
            if patient.patientType == patient_type
            and patient.scanWeek != -1
            and start_week <= patient.scanWeek < end_week
        )
        counts.append(float(count))
    return counts


def estimate_control_coefficients(X_values, YE_values, YU_values):
    """
    c = Cov(X,Y) / Var(Y), using batch observations.
    """
    X = np.array(X_values, dtype=float)
    YE = np.array(YE_values, dtype=float)
    YU = np.array(YU_values, dtype=float)

    var_YE = float(np.var(YE, ddof=1)) if len(YE) > 1 else 0.0
    var_YU = float(np.var(YU, ddof=1)) if len(YU) > 1 else 0.0

    c_E = 0.0 if var_YE == 0 else float(np.cov(X, YE, ddof=1)[0, 1] / var_YE)
    c_U = 0.0 if var_YU == 0 else float(np.cov(X, YU, ddof=1)[0, 1] / var_YU)

    return c_E, c_U


def apply_control_variates(X_values, YE_values, YU_values, v_E, v_U, c_E, c_U):
    X = np.array(X_values, dtype=float)
    YE = np.array(YE_values, dtype=float)
    YU = np.array(YU_values, dtype=float)
    return X - c_E * (YE - v_E) - c_U * (YU - v_U)


def estimate_batch_length(design):
    """
    Pilot run to estimate L_ac and then M = 5 * L_ac.
    """
    if FORCE_M is not None:
        return FORCE_M, None

    total_pilot_weeks = WARMUP_WEEKS + PILOT_WEEKS
    sim_pilot = run_simulation(design, BASE_SEED, total_pilot_weeks, use_antithetic=False)
    pilot_series = get_weekly_objective_values(sim_pilot, total_pilot_weeks)
    lag_ac = find_autocorr_lag(pilot_series)
    M = 5 * lag_ac
    return M, lag_ac


def run_one_replication(design, seed, M):
    """
    Runs one paired normal/antithetic simulation replication and returns batch-level
    objective values with variance reduction.

    Methods combined here:
    1. Batch means
    2. Common random numbers across designs through the same seed
    3. Antithetic variates
    4. Control variates based on scanned elective and urgent patients per batch
    """
    run_weeks = M * BATCHES_PER_REPLICATION
    total_weeks = WARMUP_WEEKS + run_weeks

    sim_normal = run_simulation(design, seed, total_weeks, use_antithetic=False)
    sim_anti = run_simulation(design, seed, total_weeks, use_antithetic=True)

    Xn_weekly = get_weekly_objective_values(sim_normal, total_weeks)
    Xa_weekly = get_weekly_objective_values(sim_anti, total_weeks)

    needed = M * BATCHES_PER_REPLICATION
    if len(Xn_weekly) < needed:
        Xn_weekly += [Xn_weekly[-1]] * (needed - len(Xn_weekly))
    if len(Xa_weekly) < needed:
        Xa_weekly += [Xa_weekly[-1]] * (needed - len(Xa_weekly))

    X_normal = compute_batch_means(Xn_weekly, M, BATCHES_PER_REPLICATION)
    X_anti = compute_batch_means(Xa_weekly, M, BATCHES_PER_REPLICATION)
    X_av = [(normal + anti) / 2.0 for normal, anti in zip(X_normal, X_anti)]

    YE_n = count_scanned_by_batch(sim_normal, M, BATCHES_PER_REPLICATION, patient_type=1)
    YU_n = count_scanned_by_batch(sim_normal, M, BATCHES_PER_REPLICATION, patient_type=2)
    YE_a = count_scanned_by_batch(sim_anti, M, BATCHES_PER_REPLICATION, patient_type=1)
    YU_a = count_scanned_by_batch(sim_anti, M, BATCHES_PER_REPLICATION, patient_type=2)

    YE_av = [(normal + anti) / 2.0 for normal, anti in zip(YE_n, YE_a)]
    YU_av = [(normal + anti) / 2.0 for normal, anti in zip(YU_n, YU_a)]

    v_E = 5 * M * sim_normal.lambdaElective
    v_U = M * (4 * sim_normal.lambdaUrgent[0] + 2 * sim_normal.lambdaUrgent[1])

    c_E_raw, c_U_raw = estimate_control_coefficients(X_normal, YE_n, YU_n)
    X_cv = apply_control_variates(X_normal, YE_n, YU_n, v_E, v_U, c_E_raw, c_U_raw)

    c_E_av, c_U_av = estimate_control_coefficients(X_av, YE_av, YU_av)
    X_combined = apply_control_variates(X_av, YE_av, YU_av, v_E, v_U, c_E_av, c_U_av)

    X_combined = [float(value) for value in X_combined]
    X_cv = [float(value) for value in X_cv]

    return {
        "objective": safe_mean(X_combined),
        "batch_means": X_combined,
        "raw_batch_means": X_normal,
        "anti_batch_means": X_anti,
        "antithetic_batch_means": X_av,
        "control_batch_means": X_cv,
        "YE_av": YE_av,
        "YU_av": YU_av,
        "c_E_raw": float(c_E_raw),
        "c_U_raw": float(c_U_raw),
        "c_E_av": float(c_E_av),
        "c_U_av": float(c_U_av),
        "v_E": float(v_E),
        "v_U": float(v_U),
        "red_av": var_reduction_pct(X_av, X_normal),
        "red_cv": var_reduction_pct(X_cv, X_normal),
        "red_comb": var_reduction_pct(X_combined, X_normal),
        "M": M,
        "run_weeks": run_weeks,
        "total_weeks": total_weeks,
    }


def build_pairwise_names(designs):
    """Geeft alle unieke paren: (a1, a2), (a1, a3), ..., (a3, a4)."""
    pairs = []
    for i in range(len(designs)):
        for j in range(i + 1, len(designs)):
            pairs.append((designs[i]["name"], designs[j]["name"]))
    return pairs


def summarize_difference(values, individual_confidence):
    """
    Vat de paarsgewijze verschillen samen met paired-t CI.

    individual_confidence: het PER-INTERVAL confidence niveau, dus reeds
    Bonferroni-gecorrigeerd t.o.v. de overall confidence.
    """
    mean = safe_mean(values)
    std = safe_std(values)
    half_width = ci_half(values, individual_confidence)
    ci_lower = mean - half_width
    ci_upper = mean + half_width
    significant = (ci_lower > 0.0) or (ci_upper < 0.0)
    return {
        "mean": mean,
        "std": std,
        "ci_half_width": half_width,
        "ci_lower": ci_lower,
        "ci_upper": ci_upper,
        "significant": significant,
    }


def style_header(cell, text):
    cell.value = text
    cell.font = WHITE_FONT
    cell.fill = BLUE_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_value(cell, value, number_format=None, fill=None, bold=False, align=RIGHT):
    cell.value = value
    cell.font = BOLD_FONT if bold else REG_FONT
    cell.alignment = align
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill


def write_designs_sheet(wb, individual_confidence, n_pairs, batch_info_by_design):
    ws = wb.create_sheet("Designs", 0)

    widths = [12, 18, 14, 14, 32, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "All Pairwise Comparison — Design Settings"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    headers = ["Design", "Urgent slots", "Strategy", "Rule", "Input file", "Batch info"]
    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    for row_idx, design in enumerate(DESIGNS, start=4):
        info = batch_info_by_design[design["name"]]
        values = [
            design["name"],
            design["urgent_slots"],
            design["strategy"],
            design["rule"],
            get_input_file(design),
            f"L_ac={info['lag_ac']}, M={info['M']}, run weeks={info['run_weeks']}",
        ]
        for col, value in enumerate(values, start=1):
            style_value(
                ws.cell(row=row_idx, column=col),
                value,
                number_format="#,##0" if col in [2, 3, 4] else None,
                bold=col == 1,
                align=CENTER if col <= 4 else LEFT,
            )

    parameter_start = 4 + len(DESIGNS) + 2
    ws.merge_cells(start_row=parameter_start, start_column=1, end_row=parameter_start, end_column=6)
    style_value(ws.cell(row=parameter_start, column=1), "Simulation Parameters", fill=LIGHT_FILL, bold=True, align=CENTER)

    parameters = [
        ("Warmup weeks", WARMUP_WEEKS),
        ("Long trajectories", N_REPLICATIONS),
        ("Batches per trajectory", BATCHES_PER_REPLICATION),
        ("Batch length", "Estimated per design: M = 5 * L_ac"),
        ("Pilot weeks after warmup", PILOT_WEEKS),
        ("Base seed", BASE_SEED),
        ("Variance reduction", "CRN + antithetic variates + control variates"),
        ("Overall confidence", f"{OVERALL_CONFIDENCE:.2%}"),
        ("Number of pairs (c)", n_pairs),
        ("Per-interval confidence (Bonferroni)", f"{individual_confidence:.4%}"),
    ]

    for row_idx, (label, value) in enumerate(parameters, start=parameter_start + 1):
        style_value(ws.cell(row=row_idx, column=1), label, bold=True, align=LEFT)
        style_value(ws.cell(row=row_idx, column=2), value, number_format="#,##0" if isinstance(value, int) else None)
        for col in range(3, 7):
            style_value(ws.cell(row=row_idx, column=col), "")

    ws.freeze_panes = "A4"


def write_batch_details_sheet(wb, batch_rows, pair_names):
    ws = wb.create_sheet("Batch mean details", 2)

    headers = ["Replication", "Batch"]
    headers += [f"J({design['name']}) combined" for design in DESIGNS]
    headers += [f"J({design['name']}) raw" for design in DESIGNS]
    headers += [f"J({design['name']}) antithetic" for design in DESIGNS]
    headers += [f"J({design['name']}) control" for design in DESIGNS]
    headers += [f"J({left}) - J({right}) combined" for left, right in pair_names]

    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "Batch Mean Observations from the Long Steady-State Trajectory"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    for row_idx, row in enumerate(batch_rows, start=4):
        zebra = GREY_FILL if row_idx % 2 == 0 else None
        values = [row["replication"], row["batch"]]
        values += [row["batch_objectives"][design["name"]] for design in DESIGNS]
        values += [row["raw_objectives"][design["name"]] for design in DESIGNS]
        values += [row["antithetic_objectives"][design["name"]] for design in DESIGNS]
        values += [row["control_objectives"][design["name"]] for design in DESIGNS]
        values += [row["batch_differences"][(left, right)] for left, right in pair_names]

        for col, value in enumerate(values, start=1):
            style_value(
                ws.cell(row=row_idx, column=col),
                value,
                number_format="#,##0" if col in [1, 2] else "#,##0.00000",
                fill=zebra,
                bold=col in [1, 2],
                align=CENTER if col in [1, 2] else RIGHT,
            )

    avg_row = 4 + len(batch_rows)
    style_value(ws.cell(row=avg_row, column=1), "AVG", fill=BLUE_FILL, bold=True, align=CENTER)
    ws.cell(row=avg_row, column=1).font = WHITE_FONT
    style_value(ws.cell(row=avg_row, column=2), "", fill=BLUE_FILL, bold=True, align=CENTER)
    ws.cell(row=avg_row, column=2).font = WHITE_FONT

    for col in range(3, len(headers) + 1):
        column_values = [ws.cell(row=r, column=col).value for r in range(4, avg_row)]
        style_value(
            ws.cell(row=avg_row, column=col),
            safe_mean(column_values),
            number_format="#,##0.00000",
            fill=LIGHT_FILL,
            bold=True,
        )

    ws.freeze_panes = "A4"


# New: variance reduction diagnostics sheet
def write_variance_reduction_sheet(wb, vr_rows):
    ws = wb.create_sheet("Variance reduction", 3)

    headers = [
        "Replication", "Design", "M", "v_E", "v_U",
        "c_E raw", "c_U raw", "c_E combined", "c_U combined",
        "Antithetic reduction", "Control reduction", "Combined reduction",
    ]

    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "Variance Reduction Diagnostics"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    for row_idx, row in enumerate(vr_rows, start=4):
        zebra = GREY_FILL if row_idx % 2 == 0 else None
        values = [
            row["replication"],
            row["design"],
            row["M"],
            row["v_E"],
            row["v_U"],
            row["c_E_raw"],
            row["c_U_raw"],
            row["c_E_av"],
            row["c_U_av"],
            row["red_av"] / 100.0,
            row["red_cv"] / 100.0,
            row["red_comb"] / 100.0,
        ]

        for col, value in enumerate(values, start=1):
            style_value(
                ws.cell(row=row_idx, column=col),
                value,
                number_format="#,##0" if col in [1, 3] else "0.00%" if col in [10, 11, 12] else "#,##0.000000",
                fill=zebra,
                bold=col in [1, 2],
                align=CENTER if col in [1, 2, 3] else RIGHT,
            )

    ws.freeze_panes = "A4"


def write_replications_sheet(wb, batch_rows, pair_names):
    ws = wb.create_sheet("Pairwise replications", 1)

    headers = ["Batch"]
    headers += [f"J({design['name']})" for design in DESIGNS]
    headers += [f"J({left}) - J({right})" for left, right in pair_names]

    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1)
    title.value = "Batch Mean Objective Values and Pairwise Differences"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    for row_idx, row in enumerate(batch_rows, start=4):
        zebra = GREY_FILL if row_idx % 2 == 0 else None
        values = [row["batch"]]
        values += [row["batch_objectives"][design["name"]] for design in DESIGNS]
        values += [row["batch_differences"][(left, right)] for left, right in pair_names]

        for col, value in enumerate(values, start=1):
            style_value(
                ws.cell(row=row_idx, column=col),
                value,
                number_format="#,##0" if col == 1 else "#,##0.00000",
                fill=zebra,
                bold=col == 1,
                align=CENTER if col == 1 else RIGHT,
            )

    avg_row = 4 + len(batch_rows)
    style_value(ws.cell(row=avg_row, column=1), "AVG", fill=BLUE_FILL, bold=True, align=CENTER)
    ws.cell(row=avg_row, column=1).font = WHITE_FONT

    for col in range(2, len(headers) + 1):
        column_values = [ws.cell(row=r, column=col).value for r in range(4, avg_row)]
        style_value(
            ws.cell(row=avg_row, column=col),
            safe_mean(column_values),
            number_format="#,##0.00000",
            fill=LIGHT_FILL,
            bold=True,
        )

    ws.freeze_panes = "A4"


def write_summary_sheet(wb, difference_values, individual_confidence):
    ws = wb.create_sheet("Summary", 4)

    widths = [24, 16, 16, 18, 16, 16, 18, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = (
        f"Summary of Pairwise CIs (per-interval {individual_confidence:.2%}, "
        f"overall {OVERALL_CONFIDENCE:.0%} via Bonferroni)"
    )
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER
    ws.merge_cells("A2:H2")
    explanation = ws["A2"]
    explanation.value = (
        "Each design is simulated as one long steady-state trajectory. "
        "After warmup deletion, the trajectory is divided into batches. "
        "The batch means are treated as the observations for the pairwise comparisons."
    )
    explanation.font = REG_FONT
    explanation.alignment = LEFT
    explanation.border = THIN_BORDER

    headers = [
        "Comparison",
        "Mean difference",
        "Std dev",
        "CI half-width",
        "CI lower",
        "CI upper",
        "Significant?",
        "Interpretation",
    ]
    for col, header in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=col), header)

    for row_idx, ((left, right), values) in enumerate(difference_values.items(), start=4):
        summary = summarize_difference(values, individual_confidence)

        if summary["ci_upper"] < 0:
            interpretation = f"{left} has lower J than {right}"
        elif summary["ci_lower"] > 0:
            interpretation = f"{right} has lower J than {left}"
        else:
            interpretation = "No clear difference"

        row_values = [
            f"J({left}) - J({right})",
            summary["mean"],
            summary["std"],
            summary["ci_half_width"],
            summary["ci_lower"],
            summary["ci_upper"],
            "YES" if summary["significant"] else "NO",
            interpretation,
        ]

        for col, value in enumerate(row_values, start=1):
            fill = None
            if col == 7:
                fill = GREEN_FILL if value == "YES" else RED_FILL
            style_value(
                ws.cell(row=row_idx, column=col),
                value,
                number_format="#,##0.00000" if col in [2, 3, 4, 5, 6] else None,
                fill=fill,
                bold=col in [1, 7],
                align=LEFT if col in [1, 8] else CENTER if col == 7 else RIGHT,
            )

    ws.freeze_panes = "A4"


def main():
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    pair_names = build_pairwise_names(DESIGNS)
    n_pairs = len(pair_names)

    # Bonferroni: per-interval confidence = 1 - alpha/c.
    alpha = 1.0 - OVERALL_CONFIDENCE
    individual_confidence = 1.0 - alpha / n_pairs

    replication_rows = []
    batch_rows = []
    vr_rows = []
    difference_values = {pair: [] for pair in pair_names}

    batch_info_by_design = {}
    for design in DESIGNS:
        M, lag_ac = estimate_batch_length(design)
        batch_info_by_design[design["name"]] = {
            "M": M,
            "lag_ac": lag_ac,
            "run_weeks": M * BATCHES_PER_REPLICATION,
            "total_weeks": WARMUP_WEEKS + M * BATCHES_PER_REPLICATION,
        }

    print("\nAll pairwise comparison")
    print("=" * 70)
    print(f"Long trajectories:     {N_REPLICATIONS}")
    print(f"Warmup weeks:          {WARMUP_WEEKS}")
    print("Run weeks after warmup: estimated per design as M * batches")
    print(f"Batches per trajectory:{BATCHES_PER_REPLICATION}")
    print("Batch length:          estimated per design with M = 5 * L_ac")
    print(f"Number of pairs (c):   {n_pairs}")
    print(f"Overall confidence:    {OVERALL_CONFIDENCE:.2%}")
    print(f"Per-interval (Bonf.):  {individual_confidence:.4%}")
    print("\nDesigns:")
    for design in DESIGNS:
        print(
            f"  {design['name']}: urgent_slots={design['urgent_slots']}, "
            f"strategy={design['strategy']}, rule={design['rule']}, "
            f"input={get_input_file(design)}"
        )
        info = batch_info_by_design[design["name"]]
        print(
            f"       autocorr lag L_ac={info['lag_ac']}, "
            f"batch length M={info['M']}, "
            f"run weeks={info['run_weeks']}"
        )

    # One single long trajectory.
    # The batches inside the trajectory are used as the observations.
    for replication in range(1, N_REPLICATIONS + 1):
        seed = BASE_SEED + replication
        objectives = {}

        print(f"\nReplication {replication}/{N_REPLICATIONS} | seed={seed}")

        batch_means_by_design = {}
        raw_means_by_design = {}
        antithetic_means_by_design = {}
        control_means_by_design = {}

        for design in DESIGNS:
            info = batch_info_by_design[design["name"]]
            result = run_one_replication(design, seed, info["M"])
            objective_value = result["objective"]
            batch_means_by_design[design["name"]] = result["batch_means"]
            raw_means_by_design[design["name"]] = result["raw_batch_means"]
            antithetic_means_by_design[design["name"]] = result["antithetic_batch_means"]
            control_means_by_design[design["name"]] = result["control_batch_means"]
            objectives[design["name"]] = objective_value
            print(
                f"  J({design['name']}) = {objective_value:.5f} "
                f"based on {BATCHES_PER_REPLICATION} batch means with M={info['M']}"
            )
            vr_rows.append({
                "replication": replication,
                "design": design["name"],
                "M": result["M"],
                "v_E": result["v_E"],
                "v_U": result["v_U"],
                "c_E_raw": result["c_E_raw"],
                "c_U_raw": result["c_U_raw"],
                "c_E_av": result["c_E_av"],
                "c_U_av": result["c_U_av"],
                "red_av": result["red_av"],
                "red_cv": result["red_cv"],
                "red_comb": result["red_comb"],
            })

        differences = {}
        for left, right in pair_names:
            diff = objectives[left] - objectives[right]
            differences[(left, right)] = diff
            print(f"  J({left}) - J({right}) = {diff:.5f}")

        for batch_idx in range(BATCHES_PER_REPLICATION):
            batch_objectives = {
                design["name"]: batch_means_by_design[design["name"]][batch_idx]
                for design in DESIGNS
            }
            raw_objectives = {
                design["name"]: raw_means_by_design[design["name"]][batch_idx]
                for design in DESIGNS
            }
            antithetic_objectives = {
                design["name"]: antithetic_means_by_design[design["name"]][batch_idx]
                for design in DESIGNS
            }
            control_objectives = {
                design["name"]: control_means_by_design[design["name"]][batch_idx]
                for design in DESIGNS
            }
            batch_differences = {}
            for left, right in pair_names:
                batch_diff = batch_objectives[left] - batch_objectives[right]
                batch_differences[(left, right)] = batch_diff
                difference_values[(left, right)].append(batch_diff)

            batch_rows.append({
                "replication": replication,
                "batch": batch_idx + 1,
                "batch_objectives": batch_objectives,
                "raw_objectives": raw_objectives,
                "antithetic_objectives": antithetic_objectives,
                "control_objectives": control_objectives,
                "batch_differences": batch_differences,
            })

        replication_rows.append({
            "replication": replication,
            "objectives": objectives,
            "differences": differences,
        })

    wb = Workbook()
    wb.remove(wb.active)

    write_designs_sheet(wb, individual_confidence, n_pairs, batch_info_by_design)
    write_replications_sheet(wb, batch_rows, pair_names)
    write_batch_details_sheet(wb, batch_rows, pair_names)
    write_variance_reduction_sheet(wb, vr_rows)
    write_summary_sheet(wb, difference_values, individual_confidence)

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved -> {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()