import os
import math
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from simulation import Simulation

# ============================================================
# SETTINGS
# ============================================================
DESIGNS = [
    # (urgent_slots, strategy, rule)
    (12, 3, 1),
    (16, 1, 2),
    (14, 2, 4),
    (10, 2, 3),
    (16, 3, 3),
    (14, 3, 2),
    (14, 1, 1),
    (10, 1, 2),
    
]

WARMUP_WEEKS = 50
RUN_WEEKS    = 566
TOTAL_WEEKS  = WARMUP_WEEKS + RUN_WEEKS
R            = 4
OUTPUT_EXCEL = "Big Assignment/Excel Files/combined_variance_reduction.xlsx"
INPUT_DIR    = "Big Assignment/Inputs"

# ============================================================
# Styling helpers
# ============================================================
BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
LIGHT_FILL  = PatternFill("solid", fgColor="D6E4F0")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
PURPLE_FILL = PatternFill("solid", fgColor="D7BDE2")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")

WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT  = Font(name="Arial", bold=True, size=10)
REG_FONT   = Font(name="Arial", size=10)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def style_header(cell, text, fill=None):
    cell.value     = text
    cell.font      = WHITE_FONT
    cell.fill      = fill or BLUE_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def style_label(cell, text):
    cell.value     = text
    cell.font      = BOLD_FONT
    cell.alignment = LEFT
    cell.border    = THIN_BORDER


def style_value(cell, value, fmt="#,##0.00000", fill=None):
    cell.value         = value
    cell.font          = REG_FONT
    cell.number_format = fmt
    cell.alignment     = RIGHT
    cell.border        = THIN_BORDER
    if fill:
        cell.fill = fill


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ============================================================
# Helpers
# ============================================================
def safe_avg(values):
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def run_one_replication(sim, seed, use_antithetic):
    """
    Run one replication and return post-warmup outputs.

    Antithetic variates via monkey-patch on both random.random AND random.randint,
    because helper.py's Exponential_distribution uses randint(0,1000)/1000 as its
    uniform draw — patching only random.random leaves that untouched.
      random.random   -> 1 - U
      random.randint  -> b - k  (equivalent to 1 - k/b for the 0..1000 case)
    """
    sim.resetSystem()
    random.seed(seed)

    if use_antithetic:
        _orig_random  = random.random
        _orig_randint = random.randint

        random.random = lambda: 1.0 - _orig_random()

        def _anti_randint(a, b):
            k = _orig_randint(a, b)
            return b - k   # maps k/b -> 1 - k/b

        random.randint = _anti_randint

        try:
            sim.runOneSimulation()
        finally:
            random.random  = _orig_random
            random.randint = _orig_randint
    else:
        sim.runOneSimulation()

    # Post-warmup slices
    post_el_app  = sim.movingAvgElectiveAppWT [WARMUP_WEEKS: WARMUP_WEEKS + RUN_WEEKS]
    post_ur_scan = sim.movingAvgUrgentScanWT  [WARMUP_WEEKS: WARMUP_WEEKS + RUN_WEEKS]
    post_el_scan = sim.movingAvgElectiveScanWT[WARMUP_WEEKS: WARMUP_WEEKS + RUN_WEEKS]
    post_ot      = sim.movingAvgOT            [WARMUP_WEEKS: WARMUP_WEEKS + RUN_WEEKS]

    el_app  = safe_avg(post_el_app)
    ur_scan = safe_avg(post_ur_scan)
    el_scan = safe_avg(post_el_scan)
    ot      = safe_avg(post_ot)

    xi = sim.weightEl * el_app + sim.weightUr * ur_scan

    # Control variate counts: patients scanned in post-warmup period
    ye = sum(
        1 for p in sim.patients
        if p.patientType == 1
        and p.scanWeek != -1
        and p.scanWeek >= WARMUP_WEEKS
    )
    yu = sum(
        1 for p in sim.patients
        if p.patientType == 2
        and p.scanWeek != -1
        and p.scanWeek >= WARMUP_WEEKS
    )

    return xi, ye, yu, el_app, el_scan, ur_scan, ot


# ============================================================
# Excel sheet writer
# ============================================================
def write_sheet(wb, sheet_name, results, meta):
    ws = wb.create_sheet(title=sheet_name)

    for col, w in zip("ABCDEFGHIJ", [6, 14, 14, 14, 14, 14, 22, 16, 16, 16]):
        set_col_width(ws, col, w)

    # Title
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = (f"Antithetic + Control Variates  |  Strategy {meta['strategy']}"
                   f"  |  Urgent slots {meta['urgent']}  |  Rule {meta['rule']}")
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill      = BLUE_FILL
    c.alignment = CENTER
    c.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # Summary block
    ws.merge_cells("A2:J2")
    h = ws["A2"]
    h.value = "SUMMARY"; h.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h.fill = LIGHT_FILL; h.alignment = CENTER; h.border = THIN_BORDER

    summary = [
        ("E[Y_E] = v_E",               meta["v_E"],          "E[Y_U] = v_U",          meta["v_U"]),
        ("Estimated c_E",              meta["c_E"],          "Estimated c_U",          meta["c_U"]),
        ("",                           None,                 "",                       None),
        ("Mean X̄ (raw OV)",           meta["mean_raw"],     "Std (raw)",              meta["std_raw"]),
        ("95% CI half-w (raw)",        meta["ci_raw"],       "",                       None),
        ("Mean X̄_av (antithetic)",    meta["mean_av"],      "Std (antithetic)",       meta["std_av"]),
        ("95% CI half-w (antithetic)", meta["ci_av"],        "Var.red. antithetic",    f"{meta['red_av']:.2f}%"),
        ("Mean X̄_cv (control)",       meta["mean_cv"],      "Std (control)",          meta["std_cv"]),
        ("95% CI half-w (control)",    meta["ci_cv"],        "Var.red. control",       f"{meta['red_cv']:.2f}%"),
        ("Mean X̄_combined",           meta["mean_comb"],    "Std (combined)",         meta["std_comb"]),
        ("95% CI half-w (combined)",   meta["ci_comb"],      "Var.red. combined",      f"{meta['red_comb']:.2f}%"),
    ]

    for offset, (la, va, lb, vb) in enumerate(summary, start=3):
        row = offset
        ws.row_dimensions[row].height = 16
        is_combined = "combined" in la.lower()
        fill = PURPLE_FILL if is_combined else (ORANGE_FILL if "control" in la.lower() else None)

        style_label(ws.cell(row=row, column=1), la)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        vc = ws.cell(row=row, column=4)
        if isinstance(va, float):
            style_value(vc, va, fill=fill)
        else:
            vc.value = va; vc.font = REG_FONT; vc.alignment = RIGHT; vc.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)

        ws.cell(row=row, column=6).border = THIN_BORDER  # spacer

        style_label(ws.cell(row=row, column=7), lb)
        vd = ws.cell(row=row, column=8)
        if isinstance(vb, float):
            style_value(vd, vb, fill=fill)
        else:
            vd.value = vb; vd.font = REG_FONT; vd.alignment = RIGHT; vd.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=10)

    # Detail table
    detail_start = 3 + len(summary) + 2

    ws.merge_cells(f"A{detail_start}:J{detail_start}")
    h2 = ws[f"A{detail_start}"]
    h2.value = "PER-REPLICATION DETAIL"
    h2.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h2.fill = LIGHT_FILL; h2.alignment = CENTER; h2.border = THIN_BORDER

    formula_row = detail_start + 1
    ws.merge_cells(f"A{formula_row}:J{formula_row}")
    f = ws[f"A{formula_row}"]
    f.value = ("Step 1 — Antithetic:  X_av,i = (X_normal,i + X_anti,i) / 2   "
               "  Step 2 — Control:  X_comb,i = X_av,i − c_E·(YE_av,i − v_E) − c_U·(YU_av,i − v_U)")
    f.font = Font(name="Arial", italic=True, color="595959", size=9)
    f.alignment = LEFT; f.border = THIN_BORDER

    hdr_row = formula_row + 1
    ws.row_dimensions[hdr_row].height = 20
    col_headers = [
        "Rep",
        "X_normal",
        "X_anti",
        "X_av\n(antithetic)",
        "YE_av",
        "YU_av",
        "X_combined\n(av + cv)",
        "El.AppWT",
        "Ur.ScanWT",
        "OT",
    ]
    for ci, hdr in enumerate(col_headers, start=1):
        c = ws.cell(row=hdr_row, column=ci)
        fill = PURPLE_FILL if "combined" in hdr else BLUE_FILL
        c.value = hdr; c.font = WHITE_FONT; c.fill = fill
        c.alignment = CENTER; c.border = THIN_BORDER

    for i, d in enumerate(results):
        dr = hdr_row + 1 + i
        ws.row_dimensions[dr].height = 16
        zebra = GREY_FILL if i % 2 == 0 else None

        def wc(col, val, fmt="#,##0.00000", fill=None):
            style_value(ws.cell(row=dr, column=col), val, fmt=fmt, fill=fill or zebra)

        ws.cell(row=dr, column=1).value     = i + 1
        ws.cell(row=dr, column=1).font      = BOLD_FONT
        ws.cell(row=dr, column=1).alignment = CENTER
        ws.cell(row=dr, column=1).border    = THIN_BORDER
        if zebra:
            ws.cell(row=dr, column=1).fill = zebra

        wc(2,  d["X_normal"])
        wc(3,  d["X_anti"])
        wc(4,  d["X_av"],       fill=LIGHT_FILL)
        wc(5,  d["YE_av"],      fmt="#,##0.0")
        wc(6,  d["YU_av"],      fmt="#,##0.0")
        wc(7,  d["X_combined"], fill=PURPLE_FILL)
        wc(8,  d["ElAppWT"])
        wc(9,  d["UrScanWT"])
        wc(10, d["OT"])

    # Averages footer
    avg_row = hdr_row + 1 + len(results)
    ws.row_dimensions[avg_row].height = 18
    avg_cols = [
        (None,                                           None),
        (np.mean([d["X_normal"]    for d in results]),  "#,##0.00000"),
        (np.mean([d["X_anti"]      for d in results]),  "#,##0.00000"),
        (np.mean([d["X_av"]        for d in results]),  "#,##0.00000"),
        (np.mean([d["YE_av"]       for d in results]),  "#,##0.0"),
        (np.mean([d["YU_av"]       for d in results]),  "#,##0.0"),
        (np.mean([d["X_combined"]  for d in results]),  "#,##0.00000"),
        (np.mean([d["ElAppWT"]     for d in results]),  "#,##0.00000"),
        (np.mean([d["UrScanWT"]    for d in results]),  "#,##0.00000"),
        (np.mean([d["OT"]          for d in results]),  "#,##0.00000"),
    ]
    for ci, (val, fmt) in enumerate(avg_cols, start=1):
        c = ws.cell(row=avg_row, column=ci)
        c.border = THIN_BORDER; c.font = BOLD_FONT
        if ci == 1:
            c.value = "AVG"; c.alignment = CENTER
            c.fill = BLUE_FILL; c.font = WHITE_FONT
        elif val is not None:
            c.value = val; c.number_format = fmt; c.alignment = RIGHT
            c.fill = PURPLE_FILL if ci == 7 else LIGHT_FILL

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


# ============================================================
# MAIN
# ============================================================
def main():
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for urgent, strategy, rule in DESIGNS:
        input_file = f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt"
        sheet_name = f"S{strategy}-{urgent}-R{rule}"

        print(f"\n{'='*60}")
        print(f"Design: {urgent} urgent slots | Strategy {strategy} | Rule {rule}")
        print(f"{'='*60}")

        sim = Simulation(input_file, TOTAL_WEEKS, R, rule)
        sim.setWeekSchedule()

        # Known expected values over post-warmup horizon
        v_E = 5 * RUN_WEEKS * sim.lambdaElective
        v_U = RUN_WEEKS * (4 * sim.lambdaUrgent[0] + 2 * sim.lambdaUrgent[1])

        results      = []
        X_normal_all = []
        X_av_all     = []
        YE_av_all    = []
        YU_av_all    = []
        YE_n_all     = []
        YU_n_all     = []

        for r in range(R):
            # Step 1a: normal run  (seed r, U drawn as-is)
            xi_n, ye_n, yu_n, el_app_n, el_scan_n, ur_scan_n, ot_n = \
                run_one_replication(sim, seed=r, use_antithetic=False)

            # Step 1b: antithetic run  (same seed r, U replaced by 1-U)
            xi_a, ye_a, yu_a, el_app_a, el_scan_a, ur_scan_a, ot_a = \
                run_one_replication(sim, seed=r, use_antithetic=True)

            # Step 1c: paired antithetic averages
            xi_av   = (xi_n    + xi_a)    / 2.0
            ye_av   = (ye_n    + ye_a)    / 2.0
            yu_av   = (yu_n    + yu_a)    / 2.0
            el_app  = (el_app_n  + el_app_a)  / 2.0
            el_scan = (el_scan_n + el_scan_a) / 2.0
            ur_scan = (ur_scan_n + ur_scan_a) / 2.0
            ot      = (ot_n    + ot_a)    / 2.0

            X_normal_all.append(xi_n)
            X_av_all.append(xi_av)
            YE_av_all.append(ye_av)
            YU_av_all.append(yu_av)
            YE_n_all.append(ye_n)
            YU_n_all.append(yu_n)

            results.append({
                "X_normal": xi_n,
                "X_anti":   xi_a,
                "X_av":     xi_av,
                "YE_av":    ye_av,
                "YU_av":    yu_av,
                "ElAppWT":  el_app,
                "ElScanWT": el_scan,
                "UrScanWT": ur_scan,
                "OT":       ot,
            })

            print(f"  r={r:>2} | X_n={xi_n:.5f} | X_a={xi_a:.5f} "
                  f"| X_av={xi_av:.5f} | diff={abs(xi_n - xi_a):.5f} "
                  f"| YE={ye_av:.0f} | YU={yu_av:.0f}")

        # Step 2: control variates on top of antithetic paired estimates
        X_av  = np.array(X_av_all,  dtype=float)
        YE_av = np.array(YE_av_all, dtype=float)
        YU_av = np.array(YU_av_all, dtype=float)

        var_YE = float(np.var(YE_av, ddof=1))
        var_YU = float(np.var(YU_av, ddof=1))
        c_E = 0.0 if var_YE == 0 else float(np.cov(X_av, YE_av, ddof=1)[0, 1] / var_YE)
        c_U = 0.0 if var_YU == 0 else float(np.cov(X_av, YU_av, ddof=1)[0, 1] / var_YU)

        X_combined = X_av - c_E * (YE_av - v_E) - c_U * (YU_av - v_U)

        for i, d in enumerate(results):
            d["X_combined"] = float(X_combined[i])

        # ---- Statistics ----
        X_raw = np.array(X_normal_all, dtype=float)

        mean_raw  = float(np.mean(X_raw));        std_raw  = float(np.std(X_raw,        ddof=1))
        mean_av   = float(np.mean(X_av));         std_av   = float(np.std(X_av,         ddof=1))
        mean_comb = float(np.mean(X_combined));   std_comb = float(np.std(X_combined,   ddof=1))

        # control-only baseline (pure CV op de RAW normale runs, matcht control_variates.py)
        YE_n = np.array(YE_n_all, dtype=float)
        YU_n = np.array(YU_n_all, dtype=float)

        var_YE_n = float(np.var(YE_n, ddof=1))
        var_YU_n = float(np.var(YU_n, ddof=1))
        c_E_cv = 0.0 if var_YE_n == 0 else float(np.cov(X_raw, YE_n, ddof=1)[0, 1] / var_YE_n)
        c_U_cv = 0.0 if var_YU_n == 0 else float(np.cov(X_raw, YU_n, ddof=1)[0, 1] / var_YU_n)

        X_cv_only = X_raw - c_E_cv * (YE_n - v_E) - c_U_cv * (YU_n - v_U)
        mean_cvo  = float(np.mean(X_cv_only))
        std_cvo   = float(np.std(X_cv_only, ddof=1))

        ci_raw  = 1.96 * std_raw  / np.sqrt(R)
        ci_av   = 1.96 * std_av   / np.sqrt(R)
        ci_cv   = 1.96 * std_cvo  / np.sqrt(R)
        ci_comb = 1.96 * std_comb / np.sqrt(R)

        red_av   = 100.0 * (1.0 - (std_av   / std_raw) ** 2) if std_raw > 0 else 0.0
        red_cv   = 100.0 * (1.0 - (std_cvo  / std_raw) ** 2) if std_raw > 0 else 0.0
        red_comb = 100.0 * (1.0 - (std_comb / std_raw) ** 2) if std_raw > 0 else 0.0

        print(f"\n  v_E={v_E:.1f}  v_U={v_U:.1f}  c_E={c_E:.6f}  c_U={c_U:.6f}")
        print(f"  Raw        : mean={mean_raw:.5f}  std={std_raw:.5f}  CI±{ci_raw:.5f}")
        print(f"  Antithetic : mean={mean_av:.5f}  std={std_av:.5f}  CI±{ci_av:.5f}  red={red_av:.1f}%")
        print(f"  Control    : mean={mean_cvo:.5f}  std={std_cvo:.5f}  CI±{ci_cv:.5f}  red={red_cv:.1f}%")
        print(f"  Combined   : mean={mean_comb:.5f}  std={std_comb:.5f}  CI±{ci_comb:.5f}  red={red_comb:.1f}%")

        meta = dict(
            urgent=urgent, strategy=strategy, rule=rule,
            v_E=v_E, v_U=v_U, c_E=c_E, c_U=c_U,
            mean_raw=mean_raw,   std_raw=std_raw,   ci_raw=ci_raw,
            mean_av=mean_av,     std_av=std_av,     ci_av=ci_av,     red_av=red_av,
            mean_cv=mean_cvo,    std_cv=std_cvo,    ci_cv=ci_cv,     red_cv=red_cv,
            mean_comb=mean_comb, std_comb=std_comb, ci_comb=ci_comb, red_comb=red_comb,
        )

        write_sheet(wb, sheet_name, results, meta)

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel opgeslagen → {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()