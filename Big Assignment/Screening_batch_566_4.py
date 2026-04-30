import os
import math
import random
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

from simulation import Simulation

# ============================================================
# SETTINGS  — pas hier aan
# ============================================================
#
# ============================================================
# SCREENING DESIGN FACTORS
# ============================================================
# Define the low ("-") and high ("+") level for each factor here.
# Factor 1 = number of urgent slots
# Factor 2 = timing strategy for urgent slots
# Factor 3 = appointment scheduling rule
OUTPUT_EXCEL = "Big Assignment/Excel Files/Screening_batch_566_4.xlsx"

FACTOR_LEVELS = {
    "urgent_slots": {"-": 14, "+": 16},
    "strategy": {"-": 1, "+": 2},
    "rule": {"-": 1, "+": 2},
}

FACTOR_NAMES = ["urgent_slots", "strategy", "rule"]

def build_screening_designs():
    """
    Full 2^3 screening design in the standard order used on the slides:
    1: ---
    2: +--
    3: -+-
    4: ++-
    5: --+
    6: +-+
    7: -++
    8: +++
    """
    signs = [
        ("-", "-", "-"),
        ("+", "-", "-"),
        ("-", "+", "-"),
        ("+", "+", "-"),
        ("-", "-", "+"),
        ("+", "-", "+"),
        ("-", "+", "+"),
        ("+", "+", "+"),
    ]

    designs = []
    for design_point, level_signs in enumerate(signs, start=1):
        levels = dict(zip(FACTOR_NAMES, level_signs))
        designs.append({
            "design_point": design_point,
            "levels": levels,
            "urgent": FACTOR_LEVELS["urgent_slots"][levels["urgent_slots"]],
            "strategy": FACTOR_LEVELS["strategy"][levels["strategy"]],
            "rule": FACTOR_LEVELS["rule"][levels["rule"]],
        })
    return designs


DESIGNS = build_screening_designs()

WARMUP_WEEKS = 50      # warmup period (bepaald via Welch)
L            = 4      # aantal batches

# Batch lengte:
# Slide 2: "Find L such that |Cov(Xi, Xi+L)| ≈ 0, batch length = 5L"
# We schatten de autocorrelatie-lag L_ac via de post-warmup wekelijkse data
# en stellen M = 5 * L_ac. Als je een vaste M wil, zet FORCE_M op een int.
FORCE_M      = None    # None = automatisch bepalen, int = vaste batchgrootte


INPUT_DIR    = "Big Assignment/Inputs"

# Seed voor lange runs. Normal en antithetic gebruiken dezelfde seed.
BASE_SEED = 0

# ============================================================
# Styling
# ============================================================
BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
LIGHT_FILL  = PatternFill("solid", fgColor="D6E4F0")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
PURPLE_FILL = PatternFill("solid", fgColor="D7BDE2")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")
RED_FILL    = PatternFill("solid", fgColor="FCE4D6")

WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT  = Font(name="Arial", bold=True, size=10)
REG_FONT   = Font(name="Arial", size=10)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")


def sh(cell, text, fill=None):
    cell.value = text
    cell.font = WHITE_FONT
    cell.fill = fill or BLUE_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def sl(cell, text):
    cell.value = text
    cell.font = BOLD_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def sv(cell, value, fmt="#,##0.00000", fill=None):
    cell.value = value
    cell.font = REG_FONT
    cell.number_format = fmt
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def set_col_width(ws, col, w):
    ws.column_dimensions[col].width = w


# ============================================================
# Helpers
# ============================================================
def safe_avg(values):
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def safe_np_var(values):
    arr = np.array(values, dtype=float)
    if len(arr) < 2:
        return 0.0
    return float(np.var(arr, ddof=1))


def safe_np_std(values):
    var = safe_np_var(values)
    return math.sqrt(var) if var >= 0 else 0.0


def ci_half(values):
    arr = np.array(values, dtype=float)
    if len(arr) < 2:
        return 0.0
    return float(1.96 * np.std(arr, ddof=1) / math.sqrt(len(arr)))


def ci_half_t(values, alpha=0.05):
    """
    Half-width of a two-sided 100(1-alpha)% confidence interval.
    Uses t critical values for small n. Falls back to 1.96 for large n.
    """
    arr = np.array(values, dtype=float)
    n = len(arr)
    if n < 2:
        return 0.0

    df = n - 1
    # Two-sided 95% t critical values, t_{df, 0.975}
    t_crit_95 = {
        1: 12.706,
        2: 4.303,
        3: 3.182,
        4: 2.776,
        5: 2.571,
        6: 2.447,
        7: 2.365,
        8: 2.306,
        9: 2.262,
        10: 2.228,
        11: 2.201,
        12: 2.179,
        13: 2.160,
        14: 2.145,
        15: 2.131,
        16: 2.120,
        17: 2.110,
        18: 2.101,
        19: 2.093,
        20: 2.086,
        21: 2.080,
        22: 2.074,
        23: 2.069,
        24: 2.064,
        25: 2.060,
        26: 2.056,
        27: 2.052,
        28: 2.048,
        29: 2.045,
        30: 2.042,
    }
    t_crit = t_crit_95.get(df, 1.96)
    return float(t_crit * np.std(arr, ddof=1) / math.sqrt(n))


def var_reduction_pct(new_values, raw_values):
    raw_var = safe_np_var(raw_values)
    new_var = safe_np_var(new_values)
    if raw_var <= 0:
        return 0.0
    return 100.0 * (1.0 - new_var / raw_var)


def find_autocorr_lag(series):
    """
    Zoek de kleinste lag L_ac waarvoor |autocorr(lag)| < 0.005.
    Slide 2: batchlengte M = 5 * L_ac.
    Geeft minimaal 1 terug.
    """
    arr = np.array([v for v in series if math.isfinite(v)], dtype=float)
    if len(arr) < 10:
        return 1

    arr = arr - arr.mean()
    var = np.var(arr)
    if var == 0:
        return 1

    max_lag = min(len(arr) // 2, 200)
    for lag in range(1, max_lag + 1):
        ac = np.mean(arr[:-lag] * arr[lag:]) / var
        if abs(ac) < 0.01:
            return lag
    return max_lag


# ============================================================
# Simulation runners
# ============================================================
def run_long_sim(input_file, rule, total_weeks, seed=0, use_antithetic=False):
    """
    Run één lange simulatie.

    Antithetic variates via monkey-patch op random.random en random.randint.
    Jouw helper.py gebruikt waarschijnlijk randint(0,1000)/1000 als uniform draw,
    daarom patchen we ook randint.

    randint(a,b): k -> a+b-k. Voor randint(0,1000) geeft dit 1000-k,
    dus U wordt ongeveer 1-U.
    """
    sim = Simulation(input_file, total_weeks, 1, rule)
    sim.setWeekSchedule()
    sim.resetSystem()
    random.seed(seed)

    if use_antithetic:
        _orig_random = random.random
        _orig_randint = random.randint

        random.random = lambda: 1.0 - _orig_random()

        def _anti_randint(a, b):
            k = _orig_randint(a, b)
            return a + b - k

        random.randint = _anti_randint

        try:
            sim.runOneSimulation()
        finally:
            random.random = _orig_random
            random.randint = _orig_randint
    else:
        sim.runOneSimulation()

    return sim


def get_post_warmup_weekly_ov_series(sim, run_weeks=None):
    """Geeft post-warmup wekelijkse OV-reeks terug: OV_w = weightEl*ElAppWT + weightUr*UrgScanWT."""
    el_app = sim.movingAvgElectiveAppWT
    ur_scan = sim.movingAvgUrgentScanWT

    end = len(el_app) if run_weeks is None else min(len(el_app), WARMUP_WEEKS + run_weeks)
    series = []
    for w in range(WARMUP_WEEKS, end):
        ov = sim.weightEl * el_app[w] + sim.weightUr * ur_scan[w]
        if math.isfinite(ov):
            series.append(ov)
        else:
            series.append(0.0)
    return series


def get_post_warmup_metric_series(sim, attr_name, run_weeks=None):
    values = getattr(sim, attr_name)
    end = len(values) if run_weeks is None else min(len(values), WARMUP_WEEKS + run_weeks)
    out = []
    for w in range(WARMUP_WEEKS, end):
        v = values[w]
        out.append(v if math.isfinite(v) else 0.0)
    return out


def count_scanned_by_batch(sim, M, L, patient_type):
    """
    Control variate Y per batch: aantal gescande patiënten van type patient_type
    met scanWeek in de batch na warmup.
      patient_type 1 = elective
      patient_type 2 = urgent
    """
    counts = []
    for l in range(L):
        start_week = WARMUP_WEEKS + l * M
        end_week = WARMUP_WEEKS + (l + 1) * M
        cnt = sum(
            1 for p in sim.patients
            if p.patientType == patient_type
            and p.scanWeek != -1
            and start_week <= p.scanWeek < end_week
        )
        counts.append(float(cnt))
    return counts


# ============================================================
# Batch means + variance reduction
# ============================================================
def compute_batch_means(series, M, L):
    """
    Verdeelt `series` in L batches van M weken.
    Geeft lijst van L batchgemiddelden terug.
    Slide 1: X̄_l = (1/M) * Σ f(Y_k)
    """
    batch_means = []
    for l in range(L):
        batch = series[l * M : (l + 1) * M]
        batch_means.append(safe_avg(batch))
    return batch_means


def estimate_control_coefficients(X_values, YE_values, YU_values):
    """
    c = Cov(X,Y) / Var(Y).
    Hier gebruiken we batchobservaties als sample.
    """
    X = np.array(X_values, dtype=float)
    YE = np.array(YE_values, dtype=float)
    YU = np.array(YU_values, dtype=float)

    var_YE = float(np.var(YE, ddof=1)) if len(YE) > 1 else 0.0
    var_YU = float(np.var(YU, ddof=1)) if len(YU) > 1 else 0.0

    c_E = 0.0 if var_YE == 0 else float(np.cov(X, YE, ddof=1)[0, 1] / var_YE)
    c_U = 0.0 if var_YU == 0 else float(np.cov(X, YU, ddof=1)[0, 1] / var_YU)

    return c_E, c_U


def apply_control_variates(X_values, YE_values, YU_values, v_E, v_U, c_E, c_U):
    X = np.array(X_values, dtype=float)
    YE = np.array(YE_values, dtype=float)
    YU = np.array(YU_values, dtype=float)
    return X - c_E * (YE - v_E) - c_U * (YU - v_U)


def summarize(values):
    return {
        "mean": float(np.mean(values)),
        "var": safe_np_var(values),
        "std": safe_np_std(values),
        "ci": ci_half_t(values),
    }


def sign_to_code(sign):
    return 1 if sign == "+" else -1


def calculate_screening_effects(batch_results):
    """
    Calculate main effects and two-way interaction effects for every batch.
    The response for each design point is the combined antithetic + control
    objective value X_combined for that batch.
    """
    rows = []
    for batch_idx in range(L):
        responses = {}
        codes = {}

        for result in batch_results:
            dp = result["design_point"]
            responses[dp] = result["rows"][batch_idx]["X_combined"]
            codes[dp] = {
                "F1": sign_to_code(result["levels"]["urgent_slots"]),
                "F2": sign_to_code(result["levels"]["strategy"]),
                "F3": sign_to_code(result["levels"]["rule"]),
            }

        def effect(term):
            total = 0.0
            for dp, response in responses.items():
                multiplier = 1
                for factor_code in term:
                    multiplier *= codes[dp][factor_code]
                total += multiplier * response
            return total / (2 ** (3 - 1))

        row = {
            "batch": batch_idx + 1,
            "e1_urgent_slots": effect(["F1"]),
            "e2_strategy": effect(["F2"]),
            "e3_rule": effect(["F3"]),
            "e12_urgent_strategy": effect(["F1", "F2"]),
            "e13_urgent_rule": effect(["F1", "F3"]),
            "e23_strategy_rule": effect(["F2", "F3"]),
        }
        rows.append(row)

    return rows


def summarize_effects(effect_rows):
    effect_keys = [
        "e1_urgent_slots",
        "e2_strategy",
        "e3_rule",
        "e12_urgent_strategy",
        "e13_urgent_rule",
        "e23_strategy_rule",
    ]
    summary = []
    for key in effect_keys:
        values = [row[key] for row in effect_rows]
        mean = float(np.mean(values))
        half_width = ci_half_t(values)
        summary.append({
            "effect": key,
            "mean": mean,
            "std": safe_np_std(values),
            "ci_half_width": half_width,
            "ci_lower": mean - half_width,
            "ci_upper": mean + half_width,
            "significant": (mean - half_width > 0.0) or (mean + half_width < 0.0),
        })
    return summary


#
# ============================================================
# Excel sheet schrijven
# ============================================================


def write_screening_design_sheet(wb, design_results):
    ws = wb.create_sheet(title="Screening design", index=0)

    widths = [14, 16, 16, 16, 14, 14, 14] + [14] * L + [16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(widths))
    c = ws["A1"]
    c.value = "2^3 Screening Design: Factor Levels and Objective Values"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = BLUE_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

    headers = [
        "Design point",
        "F1 urgent sign",
        "F1 urgent value",
        "F2 strategy sign",
        "F2 strategy value",
        "F3 rule sign",
        "F3 rule value",
    ] + [f"Batch {i}" for i in range(1, L + 1)] + ["Average"]

    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = text
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_idx, result in enumerate(design_results, start=4):
        levels = result["levels"]
        objective_values = [d["X_combined"] for d in result["rows"]]
        values = [
            result["design_point"],
            levels["urgent_slots"],
            result["urgent"],
            levels["strategy"],
            result["strategy"],
            levels["rule"],
            result["rule"],
        ] + objective_values + [float(np.mean(objective_values))]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.font = BOLD_FONT if col == 1 else REG_FONT
            cell.alignment = CENTER if col <= 7 else RIGHT
            cell.border = THIN_BORDER
            if col >= 8:
                cell.number_format = "#,##0.00000"

    ws.freeze_panes = "A4"


def write_screening_effects_sheet(wb, effect_rows, effect_summary):
    ws = wb.create_sheet(title="Screening effects", index=1)

    for col, w in enumerate([12, 18, 18, 18, 22, 22, 22], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Main and Interaction Effects per Batch"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = BLUE_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

    headers = [
        "Batch",
        "e1 urgent slots",
        "e2 strategy",
        "e3 rule",
        "e12 urgent×strategy",
        "e13 urgent×rule",
        "e23 strategy×rule",
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = text
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    keys = [
        "batch",
        "e1_urgent_slots",
        "e2_strategy",
        "e3_rule",
        "e12_urgent_strategy",
        "e13_urgent_rule",
        "e23_strategy_rule",
    ]
    for row_idx, row in enumerate(effect_rows, start=4):
        for col, key in enumerate(keys, start=1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = row[key]
            cell.font = REG_FONT
            cell.alignment = CENTER if col == 1 else RIGHT
            cell.border = THIN_BORDER
            if col > 1:
                cell.number_format = "#,##0.00000"

    summary_start = 4 + len(effect_rows) + 2
    ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=7)
    h = ws.cell(row=summary_start, column=1)
    h.value = "95% Confidence Intervals for Effects"
    h.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h.fill = LIGHT_FILL
    h.alignment = CENTER
    h.border = THIN_BORDER

    summary_headers = ["Effect", "Mean", "Std", "95% CI half-width", "CI lower", "CI upper", "Significant?"]
    for col, text in enumerate(summary_headers, start=1):
        cell = ws.cell(row=summary_start + 1, column=col)
        cell.value = text
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    effect_labels = {
        "e1_urgent_slots": "e1 urgent slots",
        "e2_strategy": "e2 strategy",
        "e3_rule": "e3 rule",
        "e12_urgent_strategy": "e12 urgent × strategy",
        "e13_urgent_rule": "e13 urgent × rule",
        "e23_strategy_rule": "e23 strategy × rule",
    }
    for row_idx, item in enumerate(effect_summary, start=summary_start + 2):
        values = [
            effect_labels[item["effect"]],
            item["mean"],
            item["std"],
            item["ci_half_width"],
            item["ci_lower"],
            item["ci_upper"],
            "YES" if item["significant"] else "NO",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.font = BOLD_FONT if col in [1, 7] else REG_FONT
            cell.alignment = LEFT if col == 1 else RIGHT
            cell.border = THIN_BORDER
            if col in [2, 3, 4, 5, 6]:
                cell.number_format = "#,##0.00000"
            if col == 7:
                cell.alignment = CENTER
                cell.fill = GREEN_FILL if value == "YES" else RED_FILL

    ws.freeze_panes = "A4"


# ============================================================
# Screening plots sheet
# ============================================================

def write_screening_plots_sheet(wb, design_results):
    """
    Create Excel tables and line charts for:
    - all three main effects
    - all two-way interactions

    For every interaction X_i × X_j, two separate 2D charts are made by fixing
    the third factor at '-' and '+'.
    """
    ws = wb.create_sheet(title="Screening plots", index=2)

    for col in range(1, 18):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

    ws.merge_cells("A1:Q1")
    title = ws["A1"]
    title.value = "Screening Design Graphs: Main Effects and Interaction Effects"
    title.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title.fill = BLUE_FILL
    title.alignment = CENTER
    title.border = THIN_BORDER

    factor_labels = {
        "urgent_slots": "urgent slots",
        "strategy": "strategy",
        "rule": "rule",
    }

    def response_mean(result):
        return float(np.mean([row["X_combined"] for row in result["rows"]]))

    def get_level_value(factor, sign):
        return FACTOR_LEVELS[factor][sign]

    def get_average_response(filters):
        selected = []
        for result in design_results:
            keep = True
            for factor, sign in filters.items():
                if result["levels"][factor] != sign:
                    keep = False
                    break
            if keep:
                selected.append(response_mean(result))
        return float(np.mean(selected)) if selected else 0.0

    def style_table_header(row, start_col, headers):
        for offset, text in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + offset)
            cell.value = text
            cell.font = WHITE_FONT
            cell.fill = BLUE_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def style_table_value(row, start_col, values):
        for offset, value in enumerate(values):
            cell = ws.cell(row=row, column=start_col + offset)
            cell.value = value
            cell.font = REG_FONT
            cell.alignment = CENTER if offset == 0 else RIGHT
            cell.border = THIN_BORDER
            if isinstance(value, float):
                cell.number_format = "#,##0.00000"

    def add_line_chart(title_text, data_min_col, data_max_col, cat_col, min_row, max_row, anchor, show_legend=True):
        chart = LineChart()
        chart.title = title_text
        chart.y_axis.title = "Average objective value"
        chart.x_axis.title = "Factor level"
        chart.height = 8.5
        chart.width = 14.5

        # Keep Excel's default horizontal gridlines visible, as in the example screenshots.
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.delete = False
        chart.x_axis.delete = False
        chart.legend.position = "b"
        if not show_legend:
            chart.legend = None

        data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=min_row, max_row=max_row)
        cats = Reference(ws, min_col=cat_col, min_row=min_row + 1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        for series in chart.series:
            series.marker.symbol = "none"
            series.graphicalProperties.line.width = 30000

        ws.add_chart(chart, anchor)

    # ------------------------------------------------------------
    # Main effect tables + graphs
    # ------------------------------------------------------------
    main_start = 3
    ws.cell(row=main_start, column=1).value = "MAIN EFFECTS"
    ws.cell(row=main_start, column=1).font = WHITE_FONT
    ws.cell(row=main_start, column=1).fill = BLUE_FILL
    ws.cell(row=main_start, column=1).alignment = CENTER
    ws.cell(row=main_start, column=1).border = THIN_BORDER

    main_positions = [
        ("urgent_slots", 1, "A8"),
        ("strategy", 5, "E8"),
        ("rule", 9, "I8"),
    ]

    for factor, start_col, chart_anchor in main_positions:
        style_table_header(main_start + 1, start_col, ["Level", "OV", factor_labels[factor]])
        minus_value = get_level_value(factor, "-")
        plus_value = get_level_value(factor, "+")
        minus_response = get_average_response({factor: "-"})
        plus_response = get_average_response({factor: "+"})
        style_table_value(main_start + 2, start_col, ["-", minus_response, minus_value])
        style_table_value(main_start + 3, start_col, ["+", plus_response, plus_value])
        add_line_chart(
            f"Main effect {factor_labels[factor]}",
            start_col + 1,
            start_col + 1,
            start_col + 2,
            main_start + 1,
            main_start + 3,
            chart_anchor,
            show_legend=False,
        )

    # ------------------------------------------------------------
    # Interaction effect tables + graphs
    # For every pair, create two charts by fixing the third factor at '-' and '+'.
    # ------------------------------------------------------------
    interaction_start = 25
    ws.cell(row=interaction_start, column=1).value = "INTERACTION EFFECTS"
    ws.cell(row=interaction_start, column=1).font = WHITE_FONT
    ws.cell(row=interaction_start, column=1).fill = BLUE_FILL
    ws.cell(row=interaction_start, column=1).alignment = CENTER
    ws.cell(row=interaction_start, column=1).border = THIN_BORDER

    interaction_specs = [
        ("urgent_slots", "strategy", "rule", 27, 1, "A33", "E33"),
        ("urgent_slots", "rule", "strategy", 52, 1, "A58", "E58"),
        ("strategy", "rule", "urgent_slots", 77, 1, "A83", "E83"),
    ]

    for factor_x, factor_line, fixed_factor, table_row, table_col, chart_minus_anchor, chart_plus_anchor in interaction_specs:
        for fixed_sign, chart_anchor, block_offset in [("-", chart_minus_anchor, 0), ("+", chart_plus_anchor, 4)]:
            start_col = table_col + block_offset
            fixed_value = get_level_value(fixed_factor, fixed_sign)
            block_title = ws.cell(row=table_row, column=start_col)
            block_title.value = (
                f"Interaction {factor_labels[factor_x]} × {factor_labels[factor_line]} "
                f"| {factor_labels[fixed_factor]} = {fixed_sign} ({fixed_value})"
            )
            block_title.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
            block_title.fill = LIGHT_FILL
            block_title.alignment = CENTER
            block_title.border = THIN_BORDER
            ws.merge_cells(start_row=table_row, start_column=start_col, end_row=table_row, end_column=start_col + 2)

            style_table_header(
                table_row + 1,
                start_col,
                [
                    factor_labels[factor_x],
                    f"{factor_labels[factor_line]} -",
                    f"{factor_labels[factor_line]} +",
                    f"fixed {factor_labels[fixed_factor]}",
                ],
            )

            x_minus_value = get_level_value(factor_x, "-")
            x_plus_value = get_level_value(factor_x, "+")
            y_xminus_lminus = get_average_response({factor_x: "-", factor_line: "-", fixed_factor: fixed_sign})
            y_xplus_lminus = get_average_response({factor_x: "+", factor_line: "-", fixed_factor: fixed_sign})
            y_xminus_lplus = get_average_response({factor_x: "-", factor_line: "+", fixed_factor: fixed_sign})
            y_xplus_lplus = get_average_response({factor_x: "+", factor_line: "+", fixed_factor: fixed_sign})

            style_table_value(table_row + 2, start_col, [x_minus_value, y_xminus_lminus, y_xminus_lplus])
            style_table_value(table_row + 3, start_col, [x_plus_value, y_xplus_lminus, y_xplus_lplus])

            add_line_chart(
                f"{factor_labels[factor_x]} × {factor_labels[factor_line]} | {factor_labels[fixed_factor]}={fixed_sign}",
                start_col + 1,
                start_col + 2,
                start_col,
                table_row + 1,
                table_row + 3,
                chart_anchor,
                show_legend=True,
            )

    ws.freeze_panes = "A3"


def write_batch_vr_sheet(wb, sheet_name, rows, meta):
    ws = wb.create_sheet(title=sheet_name)

    widths = [7, 12, 12, 14, 14, 14, 14, 12, 12, 14, 14, 14, 14]
    for idx, w in enumerate(widths, start=1):
        set_col_width(ws, chr(64 + idx), w)

    # Title
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = (
        f"Batch Mean Method + Variance Reduction  |  Design point {meta['design_point']}"
        f"  |  Strategy {meta['strategy']}  |  Urgent slots {meta['urgent']}  |  Rule {meta['rule']}"
    )
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = BLUE_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER
    ws.row_dimensions[1].height = 24

    # Parameters
    ws.merge_cells("A2:M2")
    h = ws["A2"]
    h.value = "PARAMETERS & METHODE"
    h.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h.fill = LIGHT_FILL
    h.alignment = CENTER
    h.border = THIN_BORDER

    param_rows = [
        ("Warmup weken", meta["warmup"], "weken"),
        ("Autocorrelatie-lag L_ac", meta["lag_ac"], "weken"),
        ("Batchgrootte M (= 5·L_ac)", meta["M"], "weken per batch"),
        ("Aantal batches L", meta["L"], "batches"),
        ("Run length na warmup", meta["run_weeks"], "weken (= M × L)"),
        ("Totale simulatielengte", meta["total_weeks"], "weken (= warmup + run)"),
        ("E[Y_E] per batch = v_E", meta["v_E"], "elective scans per batch"),
        ("E[Y_U] per batch = v_U", meta["v_U"], "urgent scans per batch"),
        ("c_E raw / c_U raw", f"{meta['c_E_raw']:.8f} / {meta['c_U_raw']:.8f}", "control-only coefficients"),
        ("c_E av / c_U av", f"{meta['c_E_av']:.8f} / {meta['c_U_av']:.8f}", "combined coefficients"),
    ]

    for i, (label, value, unit) in enumerate(param_rows, start=3):
        ws.row_dimensions[i].height = 17
        sl(ws.cell(row=i, column=1), label)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

        vc = ws.cell(row=i, column=5)
        if isinstance(value, (int, float)):
            sv(vc, value, fmt="#,##0.00000" if isinstance(value, float) else "#,##0")
        else:
            vc.value = value
            vc.font = REG_FONT
            vc.alignment = RIGHT
            vc.border = THIN_BORDER
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=7)

        ws.cell(row=i, column=8).value = unit
        ws.cell(row=i, column=8).font = Font(name="Arial", italic=True, color="595959", size=9)
        ws.cell(row=i, column=8).border = THIN_BORDER
        ws.merge_cells(start_row=i, start_column=8, end_row=i, end_column=13)

    # Summary
    sum_start = 3 + len(param_rows) + 1
    ws.merge_cells(start_row=sum_start, start_column=1, end_row=sum_start, end_column=13)
    h2 = ws.cell(row=sum_start, column=1)
    h2.value = "SUMMARY STATISTIEKEN"
    h2.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h2.fill = LIGHT_FILL
    h2.alignment = CENTER
    h2.border = THIN_BORDER

    summary_rows = [
        ("Raw batch means",       meta["raw"]["mean"],      meta["raw"]["var"],      meta["raw"]["std"],      meta["raw"]["ci"],      None,                    BLUE_FILL),
        ("Antithetic",            meta["av"]["mean"],       meta["av"]["var"],       meta["av"]["std"],       meta["av"]["ci"],       meta["red_av"],          LIGHT_FILL),
        ("Control only",          meta["cv"]["mean"],       meta["cv"]["var"],       meta["cv"]["std"],       meta["cv"]["ci"],       meta["red_cv"],          ORANGE_FILL),
        ("Combined anti + ctrl",  meta["comb"]["mean"],     meta["comb"]["var"],     meta["comb"]["std"],     meta["comb"]["ci"],     meta["red_comb"],        PURPLE_FILL),
    ]

    hdr = sum_start + 1
    headers = ["Method", "Mean", "Variance S²", "Std S", "95% CI half-width", "CI lower", "CI upper", "Var. reduction", "", "", "", "", ""]
    for ci, text in enumerate(headers, start=1):
        cell = ws.cell(row=hdr, column=ci)
        cell.value = text
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for r_i, (method, mean, var, std, ci, red, fill) in enumerate(summary_rows, start=hdr + 1):
        ws.row_dimensions[r_i].height = 18
        sl(ws.cell(row=r_i, column=1), method)
        sv(ws.cell(row=r_i, column=2), mean, fill=fill if method != "Raw batch means" else None)
        sv(ws.cell(row=r_i, column=3), var, fmt="#,##0.0000000")
        sv(ws.cell(row=r_i, column=4), std)
        sv(ws.cell(row=r_i, column=5), ci, fill=ORANGE_FILL)
        sv(ws.cell(row=r_i, column=6), mean - ci, fill=GREEN_FILL)
        sv(ws.cell(row=r_i, column=7), mean + ci, fill=GREEN_FILL)
        if red is None:
            ws.cell(row=r_i, column=8).value = "baseline"
            ws.cell(row=r_i, column=8).font = REG_FONT
            ws.cell(row=r_i, column=8).alignment = RIGHT
            ws.cell(row=r_i, column=8).border = THIN_BORDER
        else:
            sv(ws.cell(row=r_i, column=8), red / 100.0, fmt="0.00%", fill=fill)
        for ci_col in range(9, 14):
            ws.cell(row=r_i, column=ci_col).border = THIN_BORDER

    # Details table
    detail_start = hdr + len(summary_rows) + 2
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=13)
    h3 = ws.cell(row=detail_start, column=1)
    h3.value = "PER-BATCH DETAIL"
    h3.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h3.fill = LIGHT_FILL
    h3.alignment = CENTER
    h3.border = THIN_BORDER

    formula_row = detail_start + 1
    ws.merge_cells(start_row=formula_row, start_column=1, end_row=formula_row, end_column=13)
    f = ws.cell(row=formula_row, column=1)
    f.value = (
        "Batch: X_l = average weekly OV in batch.  "
        "Antithetic: X_av,l = (X_normal,l + X_anti,l)/2.  "
        "Control: X_cv,l = X_normal,l − c_E(YE_l−v_E) − c_U(YU_l−v_U).  "
        "Combined: X_comb,l = X_av,l − c_E,av(YE_av,l−v_E) − c_U,av(YU_av,l−v_U)."
    )
    f.font = Font(name="Arial", italic=True, color="595959", size=9)
    f.alignment = LEFT
    f.border = THIN_BORDER

    hdr_row = formula_row + 1
    col_headers = [
        "Batch l",
        "Week start",
        "Week einde",
        "X_normal",
        "X_anti",
        "X_av\nanti",
        "YE_av",
        "YU_av",
        "X_cv\ncontrol only",
        "X_comb\nanti+ctrl",
        "El.AppWT\navg",
        "Ur.ScanWT\navg",
        "OT\navg",
    ]
    for ci, hdr_text in enumerate(col_headers, start=1):
        cell = ws.cell(row=hdr_row, column=ci)
        fill = PURPLE_FILL if "comb" in hdr_text.lower() else BLUE_FILL
        cell.value = hdr_text
        cell.font = WHITE_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 32

    for i, d in enumerate(rows):
        dr = hdr_row + 1 + i
        zebra = GREY_FILL if i % 2 == 0 else None
        ws.row_dimensions[dr].height = 18

        def wc(col, value, fmt="#,##0.00000", fill=None):
            sv(ws.cell(row=dr, column=col), value, fmt=fmt, fill=fill or zebra)

        c1 = ws.cell(row=dr, column=1)
        c1.value = d["batch"]
        c1.font = BOLD_FONT
        c1.alignment = CENTER
        c1.border = THIN_BORDER
        if zebra:
            c1.fill = zebra

        wc(2, d["week_start"], fmt="#,##0")
        wc(3, d["week_end"], fmt="#,##0")
        wc(4, d["X_normal"])
        wc(5, d["X_anti"])
        wc(6, d["X_av"], fill=LIGHT_FILL)
        wc(7, d["YE_av"], fmt="#,##0.0")
        wc(8, d["YU_av"], fmt="#,##0.0")
        wc(9, d["X_cv"], fill=ORANGE_FILL)
        wc(10, d["X_combined"], fill=PURPLE_FILL)
        wc(11, d["ElAppWT"])
        wc(12, d["UrScanWT"])
        wc(13, d["OT"])

    # Footer averages
    avg_row = hdr_row + 1 + len(rows)
    ws.row_dimensions[avg_row].height = 18
    for ci in range(1, 14):
        cell = ws.cell(row=avg_row, column=ci)
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT

    ws.cell(row=avg_row, column=1).value = "AVG"
    ws.cell(row=avg_row, column=1).fill = BLUE_FILL
    ws.cell(row=avg_row, column=1).font = WHITE_FONT
    ws.cell(row=avg_row, column=1).alignment = CENTER

    avg_map = {
        4: "X_normal",
        5: "X_anti",
        6: "X_av",
        7: "YE_av",
        8: "YU_av",
        9: "X_cv",
        10: "X_combined",
        11: "ElAppWT",
        12: "UrScanWT",
        13: "OT",
    }
    for col, key in avg_map.items():
        value = float(np.mean([d[key] for d in rows]))
        fmt = "#,##0.0" if key in ["YE_av", "YU_av"] else "#,##0.00000"
        cell = ws.cell(row=avg_row, column=col)
        cell.value = value
        cell.number_format = fmt
        cell.alignment = RIGHT
        cell.fill = PURPLE_FILL if key == "X_combined" else LIGHT_FILL

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


# ============================================================
# MAIN
# ============================================================
def main():
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    design_results = []

    for design in DESIGNS:
        urgent = design["urgent"]
        strategy = design["strategy"]
        rule = design["rule"]
        design_point = design["design_point"]
        levels = design["levels"]

        input_file = f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt"

        print(f"\n{'=' * 70}")
        print(
            f"Design point {design_point}: "
            f"F1={levels['urgent_slots']} ({urgent} urgent slots) | "
            f"F2={levels['strategy']} (Strategy {strategy}) | "
            f"F3={levels['rule']} (Rule {rule})"
        )
        print(f"{'=' * 70}")

        # ------------------------------------------------------------
        # Step 1: pilot normal run voor autocorrelatie-lag en M
        # ------------------------------------------------------------
        PILOT_WEEKS = WARMUP_WEEKS + L * 500
        sim_pilot = run_long_sim(input_file, rule, PILOT_WEEKS, seed=BASE_SEED, use_antithetic=False)
        pilot_series = get_post_warmup_weekly_ov_series(sim_pilot)

        lag_ac = find_autocorr_lag(pilot_series)
        M = FORCE_M if FORCE_M is not None else max(5 * lag_ac, 10)
        RUN_WEEKS = M * L
        TOTAL_WEEKS = WARMUP_WEEKS + RUN_WEEKS

        print(f"  Autocorrelatie-lag L_ac = {lag_ac} weken -> M = {M} weken")
        print(f"  L = {L} batches -> run length = {RUN_WEEKS} weken, total = {TOTAL_WEEKS}")

        # ------------------------------------------------------------
        # Step 2: normal en antithetic lange run met dezelfde seed
        # ------------------------------------------------------------
        sim_normal = run_long_sim(input_file, rule, TOTAL_WEEKS, seed=BASE_SEED, use_antithetic=False)
        sim_anti   = run_long_sim(input_file, rule, TOTAL_WEEKS, seed=BASE_SEED, use_antithetic=True)

        # Weekly OV series
        Xn_weekly = get_post_warmup_weekly_ov_series(sim_normal, RUN_WEEKS)
        Xa_weekly = get_post_warmup_weekly_ov_series(sim_anti, RUN_WEEKS)

        needed = M * L
        if len(Xn_weekly) < needed:
            Xn_weekly += [Xn_weekly[-1]] * (needed - len(Xn_weekly))
        if len(Xa_weekly) < needed:
            Xa_weekly += [Xa_weekly[-1]] * (needed - len(Xa_weekly))

        # Batch means for X
        X_normal = compute_batch_means(Xn_weekly, M, L)
        X_anti   = compute_batch_means(Xa_weekly, M, L)
        X_av     = [(x + y) / 2.0 for x, y in zip(X_normal, X_anti)]

        # Extra metrics for detail table, averaged antithetic pair
        el_app_n = compute_batch_means(get_post_warmup_metric_series(sim_normal, "movingAvgElectiveAppWT", RUN_WEEKS), M, L)
        el_app_a = compute_batch_means(get_post_warmup_metric_series(sim_anti,   "movingAvgElectiveAppWT", RUN_WEEKS), M, L)
        ur_sc_n  = compute_batch_means(get_post_warmup_metric_series(sim_normal, "movingAvgUrgentScanWT", RUN_WEEKS), M, L)
        ur_sc_a  = compute_batch_means(get_post_warmup_metric_series(sim_anti,   "movingAvgUrgentScanWT", RUN_WEEKS), M, L)
        ot_n     = compute_batch_means(get_post_warmup_metric_series(sim_normal, "movingAvgOT", RUN_WEEKS), M, L)
        ot_a     = compute_batch_means(get_post_warmup_metric_series(sim_anti,   "movingAvgOT", RUN_WEEKS), M, L)

        ElAppWT  = [(a + b) / 2.0 for a, b in zip(el_app_n, el_app_a)]
        UrScanWT = [(a + b) / 2.0 for a, b in zip(ur_sc_n, ur_sc_a)]
        OT       = [(a + b) / 2.0 for a, b in zip(ot_n, ot_a)]

        # ------------------------------------------------------------
        # Step 3: control variates per batch
        # ------------------------------------------------------------
        # Verwachte aantallen per batch, niet per totale run.
        v_E = 5 * M * sim_normal.lambdaElective
        v_U = M * (4 * sim_normal.lambdaUrgent[0] + 2 * sim_normal.lambdaUrgent[1])

        YE_n = count_scanned_by_batch(sim_normal, M, L, patient_type=1)
        YU_n = count_scanned_by_batch(sim_normal, M, L, patient_type=2)
        YE_a = count_scanned_by_batch(sim_anti,   M, L, patient_type=1)
        YU_a = count_scanned_by_batch(sim_anti,   M, L, patient_type=2)

        YE_av = [(a + b) / 2.0 for a, b in zip(YE_n, YE_a)]
        YU_av = [(a + b) / 2.0 for a, b in zip(YU_n, YU_a)]

        # Control-only op raw normal batch means
        c_E_raw, c_U_raw = estimate_control_coefficients(X_normal, YE_n, YU_n)
        X_cv = apply_control_variates(X_normal, YE_n, YU_n, v_E, v_U, c_E_raw, c_U_raw)

        # Combined: antithetic + control variates
        c_E_av, c_U_av = estimate_control_coefficients(X_av, YE_av, YU_av)
        X_combined = apply_control_variates(X_av, YE_av, YU_av, v_E, v_U, c_E_av, c_U_av)

        # ------------------------------------------------------------
        # Step 4: stats + Excel rows
        # ------------------------------------------------------------
        raw_stats  = summarize(X_normal)
        av_stats   = summarize(X_av)
        cv_stats   = summarize(X_cv)
        comb_stats = summarize(X_combined)

        red_av   = var_reduction_pct(X_av, X_normal)
        red_cv   = var_reduction_pct(X_cv, X_normal)
        red_comb = var_reduction_pct(X_combined, X_normal)

        rows = []
        for i in range(L):
            rows.append({
                "batch": i + 1,
                "week_start": WARMUP_WEEKS + i * M + 1,
                "week_end": WARMUP_WEEKS + (i + 1) * M,
                "X_normal": float(X_normal[i]),
                "X_anti": float(X_anti[i]),
                "X_av": float(X_av[i]),
                "YE_av": float(YE_av[i]),
                "YU_av": float(YU_av[i]),
                "X_cv": float(X_cv[i]),
                "X_combined": float(X_combined[i]),
                "ElAppWT": float(ElAppWT[i]),
                "UrScanWT": float(UrScanWT[i]),
                "OT": float(OT[i]),
            })

        print(f"  Raw        : mean={raw_stats['mean']:.5f}  CI±{raw_stats['ci']:.5f}")
        print(f"  Antithetic : mean={av_stats['mean']:.5f}  CI±{av_stats['ci']:.5f}  red={red_av:.2f}%")
        print(f"  Control    : mean={cv_stats['mean']:.5f}  CI±{cv_stats['ci']:.5f}  red={red_cv:.2f}%")
        print(f"  Combined   : mean={comb_stats['mean']:.5f}  CI±{comb_stats['ci']:.5f}  red={red_comb:.2f}%")

        meta = dict(
            design_point=design_point,
            urgent=urgent,
            strategy=strategy,
            rule=rule,
            warmup=WARMUP_WEEKS,
            lag_ac=lag_ac,
            M=M,
            L=L,
            run_weeks=RUN_WEEKS,
            total_weeks=TOTAL_WEEKS,
            v_E=float(v_E),
            v_U=float(v_U),
            c_E_raw=float(c_E_raw),
            c_U_raw=float(c_U_raw),
            c_E_av=float(c_E_av),
            c_U_av=float(c_U_av),
            raw=raw_stats,
            av=av_stats,
            cv=cv_stats,
            comb=comb_stats,
            red_av=float(red_av),
            red_cv=float(red_cv),
            red_comb=float(red_comb),
        )

        # write_batch_vr_sheet(wb, sheet_name, rows, meta)

        design_results.append({
            "design_point": design_point,
            "levels": levels,
            "urgent": urgent,
            "strategy": strategy,
            "rule": rule,
            "rows": rows,
            "meta": meta,
        })

    effect_rows = calculate_screening_effects(design_results)
    effect_summary = summarize_effects(effect_rows)

    # Keep only the screening output sheets in the Excel file.
    write_screening_design_sheet(wb, design_results)
    write_screening_effects_sheet(wb, effect_rows, effect_summary)
    write_screening_plots_sheet(wb, design_results)

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved -> {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
