import math
import os
import random

import numpy as np
import openpyxl
from openpyxl import load_workbook

from simulation import Simulation


print(os.getcwd())


CONFIGURATIONS = [
    (14, 1, 1),
    (16, 3, 3),
    (16, 1, 2),
    (14, 2, 4),
    (12, 3, 1),
    (10, 1, 2),
    (14, 3, 2),
    (10, 2, 3),
]

WARMUP_WEEKS: int = 50
RUN_WEEKS: int = 566
R: int = 20

EXCEL_PATH: str = "Big Assignment/Excel Files/replications.xlsx"
INPUT_DIR: str = "Big Assignment/Inputs"


def safe_avg(values: list[float]) -> float:
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def sheet_name_for(config: tuple[int, int, int]) -> str:
    n_slots, strategy, rule = config
    return f"S{strategy}-{n_slots}slots-R{rule}"[:31]


def run_one_replication(sim, seed: int, warmup_weeks: int, run_weeks: int, use_antithetic: bool):
    """
    Runs one simulation replication.

    If use_antithetic=True, random.random and random.randint are patched so that
    the simulation uses 1-U instead of U where possible.
    """

    sim.resetSystem()
    random.seed(seed)

    if use_antithetic:
        original_random = random.random
        original_randint = random.randint

        random.random = lambda: 1.0 - original_random()

        def anti_randint(a, b):
            k = original_randint(a, b)
            return b - k

        random.randint = anti_randint

        try:
            sim.runOneSimulation()
        finally:
            random.random = original_random
            random.randint = original_randint
    else:
        sim.runOneSimulation()

    post_el_app = sim.movingAvgElectiveAppWT[warmup_weeks: warmup_weeks + run_weeks]
    post_ur_scan = sim.movingAvgUrgentScanWT[warmup_weeks: warmup_weeks + run_weeks]
    post_el_scan = sim.movingAvgElectiveScanWT[warmup_weeks: warmup_weeks + run_weeks]
    post_ot = sim.movingAvgOT[warmup_weeks: warmup_weeks + run_weeks]

    avg_el_app = safe_avg(post_el_app)
    avg_ur_scan = safe_avg(post_ur_scan)
    avg_el_scan = safe_avg(post_el_scan)
    avg_ot = safe_avg(post_ot)

    weighted = avg_el_app * sim.weightEl + avg_ur_scan * sim.weightUr

    # Control variates: number of scanned patients after warmup
    ye = sum(
        1 for p in sim.patients
        if p.patientType == 1
        and p.scanWeek != -1
        and p.scanWeek >= warmup_weeks
    )

    yu = sum(
        1 for p in sim.patients
        if p.patientType == 2
        and p.scanWeek != -1
        and p.scanWeek >= warmup_weeks
    )

    return {
        "weighted": weighted,
        "el_app": avg_el_app,
        "ur_scan": avg_ur_scan,
        "el_scan": avg_el_scan,
        "ot": avg_ot,
        "ye": ye,
        "yu": yu,
    }


def prepare_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print("  Sheet exists — clearing old rows.")

        for row in ws.iter_rows(min_row=2, max_col=18):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    headers = [
        "Replication",

        "Avg ElAppWT normal (h)",
        "Avg UrScanWT normal (h)",
        "Avg ElScanWT normal (h)",
        "Avg OT normal (h)",
        "Weighted Obj normal",

        "Weighted Obj anti",
        "Weighted Obj antithetic avg",

        "YE avg",
        "YU avg",

        "Weighted Obj combined CV",

        "Avg ElAppWT VR (h)",
        "Avg UrScanWT VR (h)",
        "Avg ElScanWT VR (h)",
        "Avg OT VR (h)",

        "Var Red Antithetic (%)",
        "Var Red Control (%)",
        "Var Red Combined (%)",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    return ws


def write_summary(ws, start_row: int, stats: dict):
    ws.cell(row=start_row, column=1, value="SUMMARY")
    ws.cell(row=start_row + 1, column=1, value="Mean raw")
    ws.cell(row=start_row + 1, column=2, value=round(stats["mean_raw"], 6))

    ws.cell(row=start_row + 2, column=1, value="Std raw")
    ws.cell(row=start_row + 2, column=2, value=round(stats["std_raw"], 6))

    ws.cell(row=start_row + 3, column=1, value="95% CI half-width raw")
    ws.cell(row=start_row + 3, column=2, value=round(stats["ci_raw"], 6))

    ws.cell(row=start_row + 5, column=1, value="Mean antithetic")
    ws.cell(row=start_row + 5, column=2, value=round(stats["mean_av"], 6))

    ws.cell(row=start_row + 6, column=1, value="Std antithetic")
    ws.cell(row=start_row + 6, column=2, value=round(stats["std_av"], 6))

    ws.cell(row=start_row + 7, column=1, value="95% CI half-width antithetic")
    ws.cell(row=start_row + 7, column=2, value=round(stats["ci_av"], 6))

    ws.cell(row=start_row + 8, column=1, value="Var reduction antithetic (%)")
    ws.cell(row=start_row + 8, column=2, value=round(stats["red_av"], 2))

    ws.cell(row=start_row + 10, column=1, value="Mean control")
    ws.cell(row=start_row + 10, column=2, value=round(stats["mean_cv"], 6))

    ws.cell(row=start_row + 11, column=1, value="Std control")
    ws.cell(row=start_row + 11, column=2, value=round(stats["std_cv"], 6))

    ws.cell(row=start_row + 12, column=1, value="95% CI half-width control")
    ws.cell(row=start_row + 12, column=2, value=round(stats["ci_cv"], 6))

    ws.cell(row=start_row + 13, column=1, value="Var reduction control (%)")
    ws.cell(row=start_row + 13, column=2, value=round(stats["red_cv"], 2))

    ws.cell(row=start_row + 15, column=1, value="Mean combined")
    ws.cell(row=start_row + 15, column=2, value=round(stats["mean_comb"], 6))

    ws.cell(row=start_row + 16, column=1, value="Std combined")
    ws.cell(row=start_row + 16, column=2, value=round(stats["std_comb"], 6))

    ws.cell(row=start_row + 17, column=1, value="95% CI half-width combined")
    ws.cell(row=start_row + 17, column=2, value=round(stats["ci_comb"], 6))

    ws.cell(row=start_row + 18, column=1, value="Var reduction combined (%)")
    ws.cell(row=start_row + 18, column=2, value=round(stats["red_comb"], 2))

    ws.cell(row=start_row + 20, column=1, value="Expected YE")
    ws.cell(row=start_row + 20, column=2, value=round(stats["v_E"], 6))

    ws.cell(row=start_row + 21, column=1, value="Expected YU")
    ws.cell(row=start_row + 21, column=2, value=round(stats["v_U"], 6))

    ws.cell(row=start_row + 22, column=1, value="c_E combined")
    ws.cell(row=start_row + 22, column=2, value=round(stats["c_E"], 6))

    ws.cell(row=start_row + 23, column=1, value="c_U combined")
    ws.cell(row=start_row + 23, column=2, value=round(stats["c_U"], 6))


def run_all(configurations, warmup_weeks, run_weeks, R, excel_path, input_dir):
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    total_weeks = warmup_weeks + run_weeks

    for config in configurations:
        n_slots, strategy, rule = config
        filename = os.path.join(input_dir, f"input-S{strategy}-{n_slots}.txt")
        sname = sheet_name_for(config)

        print(f"\n{'=' * 60}")
        print(f"Config: {n_slots} urgent slots | Strategy {strategy} | Rule {rule}")
        print(f"  File : {filename}")
        print(f"  Sheet: {sname}")
        print(f"{'=' * 60}")

        ws = prepare_sheet(wb, sname)

        sim = Simulation(filename, total_weeks, R, rule)
        sim.setWeekSchedule()

        # Expected values for control variates
        v_E = 5 * run_weeks * sim.lambdaElective
        v_U = run_weeks * (4 * sim.lambdaUrgent[0] + 2 * sim.lambdaUrgent[1])

        results = []

        X_raw_all = []
        X_anti_all = []
        X_av_all = []

        YE_raw_all = []
        YU_raw_all = []
        YE_av_all = []
        YU_av_all = []

        for r in range(R):
            normal = run_one_replication(
                sim=sim,
                seed=r,
                warmup_weeks=warmup_weeks,
                run_weeks=run_weeks,
                use_antithetic=False,
            )

            anti = run_one_replication(
                sim=sim,
                seed=r,
                warmup_weeks=warmup_weeks,
                run_weeks=run_weeks,
                use_antithetic=True,
            )

            x_normal = normal["weighted"]
            x_anti = anti["weighted"]
            x_av = (x_normal + x_anti) / 2.0

            ye_av = (normal["ye"] + anti["ye"]) / 2.0
            yu_av = (normal["yu"] + anti["yu"]) / 2.0

            el_app_vr = (normal["el_app"] + anti["el_app"]) / 2.0
            ur_scan_vr = (normal["ur_scan"] + anti["ur_scan"]) / 2.0
            el_scan_vr = (normal["el_scan"] + anti["el_scan"]) / 2.0
            ot_vr = (normal["ot"] + anti["ot"]) / 2.0

            X_raw_all.append(x_normal)
            X_anti_all.append(x_anti)
            X_av_all.append(x_av)

            YE_raw_all.append(normal["ye"])
            YU_raw_all.append(normal["yu"])
            YE_av_all.append(ye_av)
            YU_av_all.append(yu_av)

            results.append({
                "replication": r,

                "el_app_normal": normal["el_app"],
                "ur_scan_normal": normal["ur_scan"],
                "el_scan_normal": normal["el_scan"],
                "ot_normal": normal["ot"],
                "x_normal": x_normal,

                "x_anti": x_anti,
                "x_av": x_av,

                "ye_av": ye_av,
                "yu_av": yu_av,

                "el_app_vr": el_app_vr,
                "ur_scan_vr": ur_scan_vr,
                "el_scan_vr": el_scan_vr,
                "ot_vr": ot_vr,
            })

            print(
                f"  r={r:>3} "
                f"X_normal={x_normal:.5f} "
                f"X_anti={x_anti:.5f} "
                f"X_av={x_av:.5f} "
                f"YE_av={ye_av:.1f} "
                f"YU_av={yu_av:.1f}"
            )

        # Convert to numpy arrays
        X_raw = np.array(X_raw_all, dtype=float)
        X_av = np.array(X_av_all, dtype=float)

        YE_raw = np.array(YE_raw_all, dtype=float)
        YU_raw = np.array(YU_raw_all, dtype=float)

        YE_av = np.array(YE_av_all, dtype=float)
        YU_av = np.array(YU_av_all, dtype=float)

        # -------------------------
        # Control variates on RAW
        # -------------------------
        var_YE_raw = float(np.var(YE_raw, ddof=1))
        var_YU_raw = float(np.var(YU_raw, ddof=1))

        c_E_raw = 0.0 if var_YE_raw == 0 else float(np.cov(X_raw, YE_raw, ddof=1)[0, 1] / var_YE_raw)
        c_U_raw = 0.0 if var_YU_raw == 0 else float(np.cov(X_raw, YU_raw, ddof=1)[0, 1] / var_YU_raw)

        X_cv = X_raw - c_E_raw * (YE_raw - v_E) - c_U_raw * (YU_raw - v_U)

        # -------------------------
        # Control variates on antithetic averages
        # -------------------------
        var_YE_av = float(np.var(YE_av, ddof=1))
        var_YU_av = float(np.var(YU_av, ddof=1))

        c_E = 0.0 if var_YE_av == 0 else float(np.cov(X_av, YE_av, ddof=1)[0, 1] / var_YE_av)
        c_U = 0.0 if var_YU_av == 0 else float(np.cov(X_av, YU_av, ddof=1)[0, 1] / var_YU_av)

        X_combined = X_av - c_E * (YE_av - v_E) - c_U * (YU_av - v_U)

        for i, d in enumerate(results):
            d["x_combined"] = float(X_combined[i])

        # -------------------------
        # Statistics
        # -------------------------
        mean_raw = float(np.mean(X_raw))
        std_raw = float(np.std(X_raw, ddof=1))

        mean_av = float(np.mean(X_av))
        std_av = float(np.std(X_av, ddof=1))

        mean_cv = float(np.mean(X_cv))
        std_cv = float(np.std(X_cv, ddof=1))

        mean_comb = float(np.mean(X_combined))
        std_comb = float(np.std(X_combined, ddof=1))

        ci_raw = 1.96 * std_raw / np.sqrt(R)
        ci_av = 1.96 * std_av / np.sqrt(R)
        ci_cv = 1.96 * std_cv / np.sqrt(R)
        ci_comb = 1.96 * std_comb / np.sqrt(R)

        red_av = 100.0 * (1.0 - (std_av / std_raw) ** 2) if std_raw > 0 else 0.0
        red_cv = 100.0 * (1.0 - (std_cv / std_raw) ** 2) if std_raw > 0 else 0.0
        red_comb = 100.0 * (1.0 - (std_comb / std_raw) ** 2) if std_raw > 0 else 0.0

        print()
        print(f"  Raw        : mean={mean_raw:.5f}  std={std_raw:.5f}  CI±{ci_raw:.5f}")
        print(f"  Antithetic : mean={mean_av:.5f}   std={std_av:.5f}   CI±{ci_av:.5f}   red={red_av:.2f}%")
        print(f"  Control    : mean={mean_cv:.5f}   std={std_cv:.5f}   CI±{ci_cv:.5f}   red={red_cv:.2f}%")
        print(f"  Combined   : mean={mean_comb:.5f} std={std_comb:.5f} CI±{ci_comb:.5f} red={red_comb:.2f}%")

        # -------------------------
        # Write rows to Excel
        # -------------------------
        for i, d in enumerate(results):
            row_idx = 2 + i

            ws.cell(row=row_idx, column=1, value=d["replication"])

            ws.cell(row=row_idx, column=2, value=round(d["el_app_normal"], 5))
            ws.cell(row=row_idx, column=3, value=round(d["ur_scan_normal"], 5))
            ws.cell(row=row_idx, column=4, value=round(d["el_scan_normal"], 5))
            ws.cell(row=row_idx, column=5, value=round(d["ot_normal"], 5))
            ws.cell(row=row_idx, column=6, value=round(d["x_normal"], 6))

            ws.cell(row=row_idx, column=7, value=round(d["x_anti"], 6))
            ws.cell(row=row_idx, column=8, value=round(d["x_av"], 6))

            ws.cell(row=row_idx, column=9, value=round(d["ye_av"], 2))
            ws.cell(row=row_idx, column=10, value=round(d["yu_av"], 2))

            ws.cell(row=row_idx, column=11, value=round(d["x_combined"], 6))

            ws.cell(row=row_idx, column=12, value=round(d["el_app_vr"], 5))
            ws.cell(row=row_idx, column=13, value=round(d["ur_scan_vr"], 5))
            ws.cell(row=row_idx, column=14, value=round(d["el_scan_vr"], 5))
            ws.cell(row=row_idx, column=15, value=round(d["ot_vr"], 5))

            ws.cell(row=row_idx, column=16, value=round(red_av, 2))
            ws.cell(row=row_idx, column=17, value=round(red_cv, 2))
            ws.cell(row=row_idx, column=18, value=round(red_comb, 2))

        stats = {
            "mean_raw": mean_raw,
            "std_raw": std_raw,
            "ci_raw": ci_raw,

            "mean_av": mean_av,
            "std_av": std_av,
            "ci_av": ci_av,
            "red_av": red_av,

            "mean_cv": mean_cv,
            "std_cv": std_cv,
            "ci_cv": ci_cv,
            "red_cv": red_cv,

            "mean_comb": mean_comb,
            "std_comb": std_comb,
            "ci_comb": ci_comb,
            "red_comb": red_comb,

            "v_E": v_E,
            "v_U": v_U,
            "c_E": c_E,
            "c_U": c_U,
        }

        summary_start_row = R + 4
        write_summary(ws, summary_start_row, stats)

        wb.save(excel_path)
        print(f"  Saved: {excel_path}")

    print("\ndone")


if __name__ == "__main__":
    run_all(
        configurations=CONFIGURATIONS,
        warmup_weeks=WARMUP_WEEKS,
        run_weeks=RUN_WEEKS,
        R=R,
        excel_path=EXCEL_PATH,
        input_dir=INPUT_DIR,
    )