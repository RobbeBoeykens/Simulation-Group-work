"""
OFAT Screening Design
=====================

Implements the one-factor-at-a-time screening design as described in
Experimental_design_II (Broos Maenhout).

Three experiments are run, each varying ONE factor while fixing the other
two at the baseline levels (= current hospital configuration).

    Baseline: (urgent=14, strategy=1, rule=1)

    Experiment A  -  factor: number of urgent slots   (11 levels: 10..20)
    Experiment B  -  factor: timing strategy          (3 levels: 1,2,3)
    Experiment C  -  factor: appointment rule         (4 levels: 1,2,3,4)

For each level, R replications are run with Common Random Numbers
(same seeds 0..R-1 across all configs in an experiment).
Control variates are applied to reduce variance, as in controlvariables.py.

Confidence intervals use Bonferroni correction so that the FAMILY-WISE
confidence level is 95% within each experiment.

Output:
    - Excel with one sheet per experiment (levels + mean + CI)
    - matplotlib plot per experiment showing means with Bonferroni CIs
    - CSV of all per-replication raw outputs (for traceability)
"""

import math
import os
import random
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from simulation import Simulation

# ============================================================
# SETTINGS  —  from previous decisions
# ============================================================
WARMUP_WEEKS = 50
RUN_WEEKS    = 483
TOTAL_WEEKS  = WARMUP_WEEKS + RUN_WEEKS
R            = 16

ALPHA_FAMILY = 0.05              # family-wise error rate per experiment
BASELINE     = (14, 1, 1)        # (urgent, strategy, rule) — current hospital

INPUT_DIR    = "Big Assignment/Inputs"
OUT_DIR      = "Big Assignment/Excel Files"
OUT_XLSX     = os.path.join(OUT_DIR, "screening_ofat.xlsx")
OUT_CSV      = os.path.join(OUT_DIR, "screening_ofat_raw.csv")
PLOT_DIR     = os.path.join(OUT_DIR, "plots")

# ============================================================
# Experimental design — OFAT
# ============================================================
EXPERIMENTS = {
    "A_urgent_slots": {
        "factor_name": "Number of urgent slots",
        "levels":      list(range(10, 21)),                 # 10..20 inclusive
        "configs":     [(n, BASELINE[1], BASELINE[2]) for n in range(10, 21)],
    },
    "B_strategy": {
        "factor_name": "Timing strategy",
        "levels":      [1, 2, 3],
        "configs":     [(BASELINE[0], s, BASELINE[2]) for s in (1, 2, 3)],
    },
    "C_rule": {
        "factor_name": "Appointment rule",
        "levels":      [1, 2, 3, 4],
        "configs":     [(BASELINE[0], BASELINE[1], r) for r in (1, 2, 3, 4)],
    },
}


# ============================================================
# Helpers
# ============================================================
def safe_avg(values):
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def input_path(urgent, strategy):
    return f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt"


def check_input_files(configs):
    """Fail fast if inputs are missing."""
    missing = []
    for urgent, strategy, _ in configs:
        p = input_path(urgent, strategy)
        if not os.path.exists(p):
            missing.append(p)
    if missing:
        print("ERROR — missing input files:")
        for m in missing:
            print(f"  {m}")
        raise FileNotFoundError(
            f"{len(missing)} input files missing. Generate them before screening."
        )


def run_one_config(urgent, strategy, rule):
    """
    Run R replications for one configuration.
    Returns per-replication arrays: X (OV), Y_E (elective arrivals),
    Y_U (urgent arrivals), and raw waiting-time averages.
    """
    sim = Simulation(input_path(urgent, strategy), TOTAL_WEEKS, R, rule)
    sim.setWeekSchedule()

    X, YE, YU, EL_APP, UR_SCAN = [], [], [], [], []

    for r in range(R):
        sim.resetSystem()
        random.seed(r)                       # Common Random Numbers across configs
        sim.runOneSimulation()

        post_el_app  = sim.movingAvgElectiveAppWT[WARMUP_WEEKS : WARMUP_WEEKS + RUN_WEEKS]
        post_ur_scan = sim.movingAvgUrgentScanWT [WARMUP_WEEKS : WARMUP_WEEKS + RUN_WEEKS]

        el_app  = safe_avg(post_el_app)
        ur_scan = safe_avg(post_ur_scan)
        ov      = sim.weightEl * el_app + sim.weightUr * ur_scan

        # Control variates: arrivals (post-warmup only, to match OV horizon)
        y_e = sum(1 for p in sim.patients
                  if p.patientType == 1 and p.scanWeek != -1
                  and p.scanWeek >= WARMUP_WEEKS)
        y_u = sum(1 for p in sim.patients
                  if p.patientType == 2 and p.scanWeek != -1
                  and p.scanWeek >= WARMUP_WEEKS)

        X.append(ov)
        YE.append(y_e)
        YU.append(y_u)
        EL_APP.append(el_app)
        UR_SCAN.append(ur_scan)

    # ---- Control variate correction ----
    X  = np.array(X);  YE = np.array(YE, dtype=float);  YU = np.array(YU, dtype=float)
    v_E = 5 * RUN_WEEKS * sim.lambdaElective
    v_U = RUN_WEEKS * (4 * sim.lambdaUrgent[0] + 2 * sim.lambdaUrgent[1])
    var_YE = float(np.var(YE, ddof=1))
    var_YU = float(np.var(YU, ddof=1))
    c_E = 0.0 if var_YE == 0 else float(np.cov(X, YE, ddof=1)[0, 1] / var_YE)
    c_U = 0.0 if var_YU == 0 else float(np.cov(X, YU, ddof=1)[0, 1] / var_YU)
    X_cv = X - c_E * (YE - v_E) - c_U * (YU - v_U)

    return {
        "X_raw":    X,
        "X_cv":     X_cv,
        "YE":       YE,
        "YU":       YU,
        "el_app":   np.array(EL_APP),
        "ur_scan":  np.array(UR_SCAN),
        "c_E":      c_E,
        "c_U":      c_U,
        "v_E":      v_E,
        "v_U":      v_U,
    }


def summary_stats(X_cv, c_adj):
    """
    Given corrected per-replication values X_cv, return mean, std,
    and Bonferroni-corrected half-width for one CI.
    c_adj is the number of simultaneous intervals (for Bonferroni).
    """
    n     = len(X_cv)
    mean_ = float(np.mean(X_cv))
    std_  = float(np.std(X_cv, ddof=1))
    # Bonferroni: alpha_individual = alpha_family / c_adj
    alpha_ind = ALPHA_FAMILY / c_adj
    # two-sided t-quantile
    t_val     = stats.t.ppf(1 - alpha_ind / 2, df=n - 1)
    half_w    = t_val * std_ / math.sqrt(n)
    return mean_, std_, half_w, t_val


# ============================================================
# Excel writer
# ============================================================
BLUE  = PatternFill("solid", fgColor="1F4E79")
LIGHT = PatternFill("solid", fgColor="D6E4F0")
GREEN = PatternFill("solid", fgColor="E2EFDA")
GREY  = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD       = Font(name="Arial", bold=True, size=10)
REG        = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")


def write_experiment_sheet(wb, exp_key, exp, rows, t_val):
    ws = wb.create_sheet(exp_key[:31])

    # title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"OFAT Screening — Experiment {exp_key}  ({exp['factor_name']})"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = BLUE; c.alignment = CENTER; c.border = BORDER

    # meta
    ws["A2"] = "Baseline (other factors fixed):"
    ws["A2"].font = BOLD
    ws["B2"] = f"urgent={BASELINE[0]}, strategy={BASELINE[1]}, rule={BASELINE[2]}"
    ws["A3"] = f"Bonferroni c = {len(exp['levels'])}   |   α_family = {ALPHA_FAMILY}   |   t-critical = {t_val:.3f}"
    ws["A3"].font = Font(italic=True, size=9, color="595959")

    # header row
    headers = ["Level", "Config", "Mean X̄_cv", "Std", "CI half-width",
               "CI lower", "CI upper", "Significant vs baseline?"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci)
        cell.value = h; cell.font = WHITE_BOLD; cell.fill = BLUE
        cell.alignment = CENTER; cell.border = BORDER

    # data
    baseline_row = None
    for i, row in enumerate(rows):
        rr = 6 + i
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=rr, column=ci)
            cell.value = val
            cell.font = REG; cell.border = BORDER
            cell.alignment = CENTER
            if isinstance(val, float):
                cell.number_format = "#,##0.00000"
        # highlight baseline row
        if (row[1] == str(BASELINE)):
            baseline_row = rr
            for ci in range(1, 9):
                ws.cell(row=rr, column=ci).fill = LIGHT

    # column widths
    for col, w in zip("ABCDEFGH", [8, 16, 14, 12, 14, 12, 12, 22]):
        ws.column_dimensions[col].width = w


# ============================================================
# Plotting
# ============================================================
def plot_experiment(exp_key, exp, means, half_ws, levels):
    fig, ax = plt.subplots(figsize=(8, 5))
    x = np.arange(len(levels))
    ax.errorbar(x, means, yerr=half_ws, fmt='o', capsize=5,
                color='#1F4E79', ecolor='#F4B942', elinewidth=2,
                markersize=8, markeredgecolor='black', markeredgewidth=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels([str(l) for l in levels])
    ax.set_xlabel(exp['factor_name'], fontsize=11)
    ax.set_ylabel("Objective value  (X̄_cv)", fontsize=11)
    ax.set_title(
        f"OFAT — Experiment {exp_key}\n"
        f"Bonferroni-corrected 95% CI  (c={len(levels)})",
        fontsize=12
    )
    ax.grid(True, linestyle='--', alpha=0.4)

    # highlight baseline level
    if exp_key == "A_urgent_slots":
        base_val = BASELINE[0]
    elif exp_key == "B_strategy":
        base_val = BASELINE[1]
    else:
        base_val = BASELINE[2]
    if base_val in levels:
        idx = levels.index(base_val)
        ax.axvline(idx, color='grey', linestyle=':', alpha=0.6, label='baseline')
        ax.legend(loc='best')

    plt.tight_layout()
    outpath = os.path.join(PLOT_DIR, f"{exp_key}.png")
    plt.savefig(outpath, dpi=140)
    plt.close(fig)
    return outpath


# ============================================================
# MAIN
# ============================================================
def main():
    os.makedirs(OUT_DIR,  exist_ok=True)
    os.makedirs(PLOT_DIR, exist_ok=True)

    # Pre-flight: check that all needed input files exist
    all_configs = []
    for exp in EXPERIMENTS.values():
        all_configs.extend(exp["configs"])
    check_input_files(all_configs)

    wb = Workbook()
    wb.remove(wb.active)

    # CSV header
    csv_lines = ["experiment,urgent,strategy,rule,rep,X_raw,X_cv,YE,YU,el_app,ur_scan"]

    for exp_key, exp in EXPERIMENTS.items():
        print(f"\n{'='*70}")
        print(f"  Experiment {exp_key}  —  varying: {exp['factor_name']}")
        print(f"  levels: {exp['levels']}")
        print(f"{'='*70}")

        c_adj = len(exp["levels"])           # Bonferroni count
        rows = []
        means, half_ws = [], []

        for level, cfg in zip(exp["levels"], exp["configs"]):
            urgent, strategy, rule = cfg
            print(f"\n  -> level = {level}   (config = {cfg})")

            out = run_one_config(urgent, strategy, rule)
            X_cv = out["X_cv"]

            mean_, std_, hw, tval = summary_stats(X_cv, c_adj)
            means.append(mean_)
            half_ws.append(hw)

            # store row for Excel
            rows.append([
                level,
                str(cfg),
                mean_, std_, hw,
                mean_ - hw, mean_ + hw,
                "—"         # fill in after loop (needs baseline)
            ])

            # append to CSV
            for r in range(R):
                csv_lines.append(
                    f"{exp_key},{urgent},{strategy},{rule},{r},"
                    f"{out['X_raw'][r]:.6f},{out['X_cv'][r]:.6f},"
                    f"{int(out['YE'][r])},{int(out['YU'][r])},"
                    f"{out['el_app'][r]:.6f},{out['ur_scan'][r]:.6f}"
                )

            print(f"     mean={mean_:.5f}  std={std_:.5f}  "
                  f"CI_bonf=±{hw:.5f}  c_E={out['c_E']:.3f}  c_U={out['c_U']:.3f}")

        # ---- Baseline comparison: non-overlap test ----
        # find the baseline row within this experiment
        baseline_idx = None
        for i, cfg in enumerate(exp["configs"]):
            if cfg == BASELINE:
                baseline_idx = i
                break

        if baseline_idx is not None:
            bmean = means[baseline_idx]
            bhw   = half_ws[baseline_idx]
            b_lo, b_hi = bmean - bhw, bmean + bhw
            for i, row in enumerate(rows):
                if i == baseline_idx:
                    row[-1] = "(baseline)"
                else:
                    lo, hi = row[5], row[6]
                    overlap = not (hi < b_lo or lo > b_hi)
                    row[-1] = "no (overlaps)" if overlap else "YES"
        else:
            # experiment without baseline level — skip comparison
            for row in rows:
                row[-1] = "n/a"

        # ---- Write Excel ----
        tval_row = stats.t.ppf(1 - (ALPHA_FAMILY / c_adj) / 2, df=R - 1)
        write_experiment_sheet(wb, exp_key, exp, rows, tval_row)

        # ---- Plot ----
        plot_path = plot_experiment(exp_key, exp, means, half_ws, exp["levels"])
        print(f"\n  Plot saved: {plot_path}")

    wb.save(OUT_XLSX)
    with open(OUT_CSV, "w") as f:
        f.write("\n".join(csv_lines))

    print(f"\n{'='*70}")
    print(f"  Excel:  {OUT_XLSX}")
    print(f"  CSV:    {OUT_CSV}")
    print(f"  Plots:  {PLOT_DIR}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
