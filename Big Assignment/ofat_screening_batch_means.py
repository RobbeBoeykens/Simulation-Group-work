"""
OFAT Screening Design — Batch Means Method
==========================================

Based on the batch mean methodology in Batch_mean_methodpy.

Baseline: (urgent=14, strategy=1, rule=1)

Experiment A: number of urgent slots   10..20
Experiment B: timing strategy          1..3
Experiment C: appointment rule         1..4

For every configuration:
    1. Run a long pilot simulation after warm-up to estimate autocorrelation lag L_ac.
    2. Set batch length M = 5 * L_ac, unless FORCE_M is specified.
    3. Run one long simulation of WARMUP_WEEKS + M * L weeks.
    4. Remove warm-up.
    5. Split the remaining weekly OV series into L non-overlapping batches.
    6. Use the L batch means as observations for mean, variance and Bonferroni CI.

Output:
    - Excel workbook with one sheet per OFAT experiment
    - CSV with all batch-level observations
    - plots with Bonferroni-corrected confidence intervals
"""

import math
import os
import random
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from simulation import Simulation

WARMUP_WEEKS = 50
L = 16
FORCE_M = None
ALPHA_FAMILY = 0.05
BASELINE = (14, 1, 1)

INPUT_DIR = "Big Assignment/Inputs"
OUT_DIR = "Big Assignment/Excel Files"
OUT_XLSX = os.path.join(OUT_DIR, "screening_ofat_batch_means.xlsx")
OUT_CSV = os.path.join(OUT_DIR, "screening_ofat_batch_means_raw.csv")
PLOT_DIR = os.path.join(OUT_DIR, "plots_batch_means")

EXPERIMENTS = {
    "A_urgent_slots": {
        "factor_name": "Number of urgent slots",
        "levels": list(range(10, 21)),
        "configs": [(n, BASELINE[1], BASELINE[2]) for n in range(10, 21)],
    },
    "B_strategy": {
        "factor_name": "Timing strategy",
        "levels": [1, 2, 3],
        "configs": [(BASELINE[0], s, BASELINE[2]) for s in (1, 2, 3)],
    },
    "C_rule": {
        "factor_name": "Appointment rule",
        "levels": [1, 2, 3, 4],
        "configs": [(BASELINE[0], BASELINE[1], r) for r in (1, 2, 3, 4)],
    },
}

BLUE = PatternFill("solid", fgColor="1F4E79")
LIGHT = PatternFill("solid", fgColor="D6E4F0")
ORANGE = PatternFill("solid", fgColor="F4B942")
GREY = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
REG = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def input_path(urgent, strategy):
    return f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt"


def safe_avg(values):
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def check_input_files(configs):
    missing = []
    for urgent, strategy, _ in configs:
        p = input_path(urgent, strategy)
        if not os.path.exists(p):
            missing.append(p)
    if missing:
        print("ERROR — missing input files:")
        for m in missing:
            print(f"  {m}")
        raise FileNotFoundError(f"{len(missing)} input files missing.")


def find_autocorr_lag(series):
    arr = np.array([v for v in series if math.isfinite(v)], dtype=float)
    if len(arr) < 10:
        return 1
    arr = arr - arr.mean()
    var = np.var(arr)
    if var == 0:
        return 1
    max_lag = min(len(arr) // 2, 200)
    for lag in range(1, max_lag + 1):
        ac = np.mean(arr[:-lag] * arr[lag:]) / var
        if abs(ac) < 0.005:
            return lag
    return max_lag


def run_long_sim(input_file, rule, total_weeks, seed=0):
    sim = Simulation(input_file, total_weeks, 1, rule)
    sim.setWeekSchedule()
    sim.resetSystem()
    random.seed(seed)
    sim.runOneSimulation()
    return sim


def get_post_warmup_ov_series(sim):
    series = []
    for w in range(WARMUP_WEEKS, len(sim.movingAvgElectiveAppWT)):
        el_app = sim.movingAvgElectiveAppWT[w]
        ur_scan = sim.movingAvgUrgentScanWT[w]
        ov = sim.weightEl * el_app + sim.weightUr * ur_scan
        if math.isfinite(ov):
            series.append(ov)
    return series


def compute_batch_means(series, M, L):
    batch_means = []
    for l in range(L):
        batch = series[l * M: (l + 1) * M]
        batch_means.append(safe_avg(batch))
    return np.array(batch_means, dtype=float)


def run_one_config_batch_means(urgent, strategy, rule):
    input_file = input_path(urgent, strategy)
    pilot_weeks = WARMUP_WEEKS + L * 500
    sim_pilot = run_long_sim(input_file, rule, pilot_weeks, seed=0)
    pilot_series = get_post_warmup_ov_series(sim_pilot)

    lag_ac = find_autocorr_lag(pilot_series)
    M = int(FORCE_M) if FORCE_M is not None else max(5 * lag_ac, 10)
    run_weeks = M * L
    total_weeks = WARMUP_WEEKS + run_weeks

    sim = run_long_sim(input_file, rule, total_weeks, seed=0)
    series = get_post_warmup_ov_series(sim)

    needed = M * L
    if len(series) < needed:
        if len(series) == 0:
            raise ValueError("Post-warmup series is empty. Check simulation output.")
        series = series + [series[-1]] * (needed - len(series))

    series = series[:needed]
    batch_means = compute_batch_means(series, M, L)

    return {
        "urgent": urgent,
        "strategy": strategy,
        "rule": rule,
        "lag_ac": lag_ac,
        "M": M,
        "L": L,
        "run_weeks": run_weeks,
        "total_weeks": total_weeks,
        "series": np.array(series, dtype=float),
        "batch_means": batch_means,
    }


def batch_summary(batch_means, c_adj):
    n = len(batch_means)
    mean_ = float(np.mean(batch_means))
    S2 = float(np.var(batch_means, ddof=1))
    std_ = math.sqrt(S2)
    alpha_ind = ALPHA_FAMILY / c_adj
    t_val = stats.t.ppf(1 - alpha_ind / 2, df=n - 1)
    half_w = t_val * math.sqrt(S2 / n)
    return mean_, S2, std_, half_w, t_val


def write_experiment_sheet(wb, exp_key, exp, rows, t_val):
    ws = wb.create_sheet(exp_key[:31])
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"OFAT Screening with Batch Means — {exp_key} ({exp['factor_name']})"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = BLUE
    c.alignment = CENTER
    c.border = BORDER

    ws["A2"] = "Baseline:"
    ws["A2"].font = BOLD
    ws["B2"] = f"urgent={BASELINE[0]}, strategy={BASELINE[1]}, rule={BASELINE[2]}"
    ws["A3"] = f"Warm-up = {WARMUP_WEEKS} weeks | L = {L} batches | Bonferroni c = {len(exp['levels'])} | α_family = {ALPHA_FAMILY} | t-critical = {t_val:.3f}"
    ws["A3"].font = Font(name="Arial", italic=True, size=9, color="595959")

    headers = [
        "Level", "Config", "Lag L_ac", "Batch length M", "Run weeks",
        "Mean X̄", "S²", "Std S", "CI half-width", "CI lower", "CI upper",
        "Significant vs baseline?"
    ]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci)
        cell.value = h
        cell.font = WHITE_BOLD
        cell.fill = BLUE
        cell.alignment = CENTER
        cell.border = BORDER

    baseline_row = None
    for i, row in enumerate(rows):
        rr = 6 + i
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=rr, column=ci)
            cell.value = val
            cell.font = REG
            cell.border = BORDER
            cell.alignment = CENTER if ci != 2 else LEFT
            if isinstance(val, float):
                cell.number_format = "#,##0.00000"
        if row[1] == str(BASELINE):
            baseline_row = rr
            for ci in range(1, len(headers) + 1):
                ws.cell(row=rr, column=ci).fill = LIGHT

    for col, width in zip("ABCDEFGHIJKL", [8, 16, 10, 15, 12, 12, 12, 12, 14, 12, 12, 24]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A6"


def write_detail_sheet(wb, exp_key, level, out):
    urgent, strategy, rule = out["urgent"], out["strategy"], out["rule"]
    sheet_name = f"{exp_key[:1]}_L{level}_detail"[:31]
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Batch detail — {exp_key}, level {level}, config {(urgent, strategy, rule)}"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = BLUE
    c.alignment = CENTER
    c.border = BORDER

    meta = [
        ("Warm-up weeks", WARMUP_WEEKS),
        ("Autocorrelation lag L_ac", out["lag_ac"]),
        ("Batch length M", out["M"]),
        ("Number of batches L", out["L"]),
        ("Run weeks", out["run_weeks"]),
        ("Total weeks", out["total_weeks"]),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=1).font = BOLD
        ws.cell(row=i, column=2).value = value
        ws.cell(row=i, column=2).alignment = RIGHT

    start = 11
    headers = ["Batch l", "Week start", "Week end", "X̄_l", "X̄_l - X̄", "(X̄_l - X̄)^2"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=ci)
        cell.value = h
        cell.font = WHITE_BOLD
        cell.fill = BLUE
        cell.alignment = CENTER
        cell.border = BORDER

    bms = out["batch_means"]
    xbar = float(np.mean(bms))
    for i, bm in enumerate(bms):
        rr = start + 1 + i
        week_start = WARMUP_WEEKS + i * out["M"] + 1
        week_end = WARMUP_WEEKS + (i + 1) * out["M"]
        diff = float(bm - xbar)
        values = [i + 1, week_start, week_end, float(bm), diff, diff ** 2]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=rr, column=ci)
            cell.value = val
            cell.font = REG
            cell.border = BORDER
            cell.alignment = CENTER if ci <= 3 else RIGHT
            if ci >= 4:
                cell.number_format = "#,##0.00000"
            if i % 2 == 0:
                cell.fill = GREY

    for col, width in zip("ABCDEFG", [10, 12, 12, 14, 14, 16, 12]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


def plot_experiment(exp_key, exp, means, half_ws, levels):
    fig, ax = plt.subplots(figsize=(8, 5))
    x = np.arange(len(levels))
    ax.errorbar(x, means, yerr=half_ws, fmt="o", capsize=5,
                color="#1F4E79", ecolor="#F4B942", elinewidth=2,
                markersize=8, markeredgecolor="black", markeredgewidth=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels([str(l) for l in levels])
    ax.set_xlabel(exp["factor_name"], fontsize=11)
    ax.set_ylabel("Objective value (batch mean estimate)", fontsize=11)
    ax.set_title(f"OFAT — {exp_key}\nBatch means with Bonferroni-corrected 95% CI", fontsize=12)
    ax.grid(True, linestyle="--", alpha=0.4)

    if exp_key == "A_urgent_slots":
        base_val = BASELINE[0]
    elif exp_key == "B_strategy":
        base_val = BASELINE[1]
    else:
        base_val = BASELINE[2]

    if base_val in levels:
        idx = levels.index(base_val)
        ax.axvline(idx, color="grey", linestyle=":", alpha=0.6, label="baseline")
        ax.legend(loc="best")

    plt.tight_layout()
    outpath = os.path.join(PLOT_DIR, f"{exp_key}_batch_means.png")
    plt.savefig(outpath, dpi=140)
    plt.close(fig)
    return outpath


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(PLOT_DIR, exist_ok=True)

    all_configs = []
    for exp in EXPERIMENTS.values():
        all_configs.extend(exp["configs"])
    check_input_files(all_configs)

    wb = Workbook()
    wb.remove(wb.active)

    csv_lines = [
        "experiment,level,urgent,strategy,rule,batch,M,L,lag_ac,week_start,week_end,X_batch_mean"
    ]

    for exp_key, exp in EXPERIMENTS.items():
        print(f"\n{'=' * 70}")
        print(f"Experiment {exp_key} — varying: {exp['factor_name']}")
        print(f"Levels: {exp['levels']}")
        print(f"{'=' * 70}")

        c_adj = len(exp["levels"])
        rows = []
        means = []
        half_ws = []
        tval_last = None
        outputs_for_detail = []

        for level, cfg in zip(exp["levels"], exp["configs"]):
            urgent, strategy, rule = cfg
            print(f"\n  -> level = {level}   config = {cfg}")

            out = run_one_config_batch_means(urgent, strategy, rule)
            bms = out["batch_means"]
            mean_, S2, std_, hw, tval = batch_summary(bms, c_adj)
            tval_last = tval

            means.append(mean_)
            half_ws.append(hw)
            outputs_for_detail.append((level, out))

            rows.append([
                level,
                str(cfg),
                out["lag_ac"],
                out["M"],
                out["run_weeks"],
                mean_,
                S2,
                std_,
                hw,
                mean_ - hw,
                mean_ + hw,
                "—",
            ])

            for i, bm in enumerate(bms):
                week_start = WARMUP_WEEKS + i * out["M"] + 1
                week_end = WARMUP_WEEKS + (i + 1) * out["M"]
                csv_lines.append(
                    f"{exp_key},{level},{urgent},{strategy},{rule},{i+1},"
                    f"{out['M']},{out['L']},{out['lag_ac']},{week_start},{week_end},{float(bm):.8f}"
                )

            print(
                f"     L_ac={out['lag_ac']}  M={out['M']}  run weeks={out['run_weeks']}  "
                f"mean={mean_:.5f}  S={std_:.5f}  CI_bonf=±{hw:.5f}"
            )

        baseline_idx = None
        for i, cfg in enumerate(exp["configs"]):
            if cfg == BASELINE:
                baseline_idx = i
                break

        if baseline_idx is not None:
            bmean = means[baseline_idx]
            bhw = half_ws[baseline_idx]
            b_lo, b_hi = bmean - bhw, bmean + bhw
            for i, row in enumerate(rows):
                if i == baseline_idx:
                    row[-1] = "(baseline)"
                else:
                    lo, hi = row[9], row[10]
                    overlap = not (hi < b_lo or lo > b_hi)
                    row[-1] = "no (overlaps)" if overlap else "YES"
        else:
            for row in rows:
                row[-1] = "n/a"

        write_experiment_sheet(wb, exp_key, exp, rows, tval_last)

        for level, out in outputs_for_detail:
            write_detail_sheet(wb, exp_key, level, out)

        plot_path = plot_experiment(exp_key, exp, means, half_ws, exp["levels"])
        print(f"\n  Plot saved: {plot_path}")

    wb.save(OUT_XLSX)
    with open(OUT_CSV, "w", encoding="utf-8") as f:
        f.write("\n".join(csv_lines))

    print(f"\n{'=' * 70}")
    print(f"Excel: {OUT_XLSX}")
    print(f"CSV:   {OUT_CSV}")
    print(f"Plots: {PLOT_DIR}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
