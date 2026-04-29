import math
import numpy as np
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from simulation import Simulation

# ============================================================
# SETTINGS
# ============================================================
DESIGNS = [
    # (urgent_slots, strategy, rule)
    (14, 1, 1),
    (16, 3, 3),
    (16, 1, 2),
    (14, 2, 4),
    (12, 3, 1),
    (10, 1, 2),
    (14, 3, 2),
    (10, 2, 3),
]

WARMUP_WEEKS = 50
RUN_WEEKS    = 483
TOTAL_WEEKS  = WARMUP_WEEKS + RUN_WEEKS
R            = 16
OUTPUT_EXCEL = "Big Assignment/Excel Files/control_variates.xlsx"
INPUT_DIR    = "Big Assignment/Inputs"

# ============================================================
# Styling helpers
# ============================================================
BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
LIGHT_FILL  = PatternFill("solid", fgColor="D6E4F0")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")

WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT  = Font(name="Arial", bold=True, size=10)
REG_FONT   = Font(name="Arial", size=10)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def style_header(cell, text, fill=None):
    cell.value     = text
    cell.font      = WHITE_FONT
    cell.fill      = fill or BLUE_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def style_label(cell, text):
    cell.value     = text
    cell.font      = BOLD_FONT
    cell.alignment = LEFT
    cell.border    = THIN_BORDER


def style_value(cell, value, fmt="#,##0.00000", fill=None):
    cell.value         = value
    cell.font          = REG_FONT
    cell.number_format = fmt
    cell.alignment     = RIGHT
    cell.border        = THIN_BORDER
    if fill:
        cell.fill = fill


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ============================================================
# Helpers
# ============================================================
def safe_avg(values: list) -> float:
    """Average ignoring NaN / Inf (weeks with 0 patients produce inf)."""
    valid = [v for v in values if math.isfinite(v)]
    return sum(valid) / len(valid) if valid else 0.0


def compute_replication_outputs(sim):
    """
    Returns the per-replication outputs needed for control variates.

    Xi       : weighted objective (OV) averaged over post-warmup weeks
    YE_i     : number of elective patients scheduled in post-warmup weeks
    YU_i     : number of urgent   patients scheduled in post-warmup weeks

    IMPORTANT: Y_E and Y_U are counted over the SAME post-warmup horizon
    as X_i so that E[Y_E] = v_E and E[Y_U] = v_U hold on the same time
    scale, keeping the corrected estimator unbiased.
    """
    # ---------- primary response Xi ----------
    post_el_app  = sim.movingAvgElectiveAppWT [WARMUP_WEEKS : WARMUP_WEEKS + RUN_WEEKS]
    post_ur_scan = sim.movingAvgUrgentScanWT  [WARMUP_WEEKS : WARMUP_WEEKS + RUN_WEEKS]
    post_el_scan = sim.movingAvgElectiveScanWT[WARMUP_WEEKS : WARMUP_WEEKS + RUN_WEEKS]
    post_ot      = sim.movingAvgOT            [WARMUP_WEEKS : WARMUP_WEEKS + RUN_WEEKS]

    el_app_wt  = safe_avg(post_el_app)
    ur_scan_wt = safe_avg(post_ur_scan)
    el_scan_wt = safe_avg(post_el_scan)
    ot         = safe_avg(post_ot)

    xi = sim.weightEl * el_app_wt + sim.weightUr * ur_scan_wt

    # ---------- control variates: count arrivals in post-warmup only ----------
    # scanWeek is set during schedulePatients(); patients with scanWeek == -1
    # were never scheduled (beyond horizon) — exclude them.
    ye_i = sum(
        1 for p in sim.patients
        if p.patientType == 1
        and p.scanWeek != -1
        and p.scanWeek >= WARMUP_WEEKS
    )
    yu_i = sum(
        1 for p in sim.patients
        if p.patientType == 2
        and p.scanWeek != -1
        and p.scanWeek >= WARMUP_WEEKS
    )

    return xi, ye_i, yu_i, el_app_wt, el_scan_wt, ur_scan_wt, ot


# ============================================================
# Write one design sheet
# ============================================================
def write_design_sheet(wb, sheet_name, results, meta):
    ws = wb.create_sheet(title=sheet_name)

    for col, w in zip("ABCDEFG", [6, 16, 16, 16, 22, 16, 16]):
        set_col_width(ws, col, w)

    # ---- title ----
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = (f"Control Variates  |  Strategy {meta['strategy']}"
                   f"  |  Urgent slots {meta['urgent']}  |  Rule {meta['rule']}")
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill      = BLUE_FILL
    c.alignment = CENTER
    c.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # ---- summary block ----
    ws.merge_cells("A2:G2")
    h = ws["A2"]
    h.value     = "SUMMARY"
    h.font      = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h.fill      = LIGHT_FILL
    h.alignment = CENTER
    h.border    = THIN_BORDER

    summary = [
        # (left label,            left value,          right label,         right value)
        ("E[Y_E]  =  v_E",       meta["v_E"],         "E[Y_U]  =  v_U",    meta["v_U"]),
        ("Estimated c_E",         meta["c_E"],         "Estimated c_U",      meta["c_U"]),
        ("",                      None,                "",                   None),
        ("Mean X̄  (raw OV)",     meta["mean_raw"],    "Std (raw)",          meta["std_raw"]),
        ("95% CI half-w (raw)",   meta["ci_raw"],      "",                   None),
        ("Mean X̄_cv (corrected)", meta["mean_cv"],     "Std (cv)",           meta["std_cv"]),
        ("95% CI half-w (cv)",    meta["ci_cv"],       "Var. reduction",     f"{meta['reduction_pct']:.2f}%"),
    ]

    for offset, (la, va, lb, vb) in enumerate(summary, start=3):
        row = offset
        ws.row_dimensions[row].height = 16

        style_label(ws.cell(row=row, column=1), la)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        vc = ws.cell(row=row, column=3)
        fill = ORANGE_FILL if "corrected" in la.lower() or "cv" in la.lower() else None
        if isinstance(va, float):
            style_value(vc, va, fill=fill)
        else:
            vc.value = va; vc.font = REG_FONT; vc.alignment = RIGHT; vc.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

        ws.cell(row=row, column=5).border = THIN_BORDER  # spacer

        style_label(ws.cell(row=row, column=6), lb)
        vd = ws.cell(row=row, column=7)
        fill2 = ORANGE_FILL if "corrected" in lb.lower() or "X̄_cv" in lb else None
        if isinstance(vb, float):
            style_value(vd, vb, fill=fill2)
        else:
            vd.value = vb; vd.font = REG_FONT; vd.alignment = RIGHT; vd.border = THIN_BORDER

    # ---- detail table ----
    detail_start = 3 + len(summary) + 2

    ws.merge_cells(f"A{detail_start}:G{detail_start}")
    h2 = ws[f"A{detail_start}"]
    h2.value     = "PER-REPLICATION DETAIL"
    h2.font      = Font(name="Arial", bold=True, color="1F4E79", size=10)
    h2.fill      = LIGHT_FILL
    h2.alignment = CENTER
    h2.border    = THIN_BORDER

    formula_row = detail_start + 1
    ws.merge_cells(f"A{formula_row}:G{formula_row}")
    f = ws[f"A{formula_row}"]
    f.value     = ("X_cv,i  =  Xᵢ  −  c_E·(YE_i − v_E)  −  c_U·(YU_i − v_U)"
                   "        [X̄_cv = X̄ − c·(Ȳ − v)]")
    f.font      = Font(name="Arial", italic=True, color="595959", size=9)
    f.alignment = LEFT
    f.border    = THIN_BORDER

    hdr_row = formula_row + 1
    ws.row_dimensions[hdr_row].height = 20
    col_headers = ["Rep", "Xᵢ (OV raw)", "YE_i (el.arr)", "YU_i (ur.arr)",
                   "Xᵢ_cv (corrected)", "El.AppWT", "Ur.ScanWT"]
    col_fills = [BLUE_FILL] * 4 + [PatternFill("solid", fgColor="1A5276")] + [BLUE_FILL] * 2
    for ci, (hdr, fl) in enumerate(zip(col_headers, col_fills), start=1):
        c = ws.cell(row=hdr_row, column=ci)
        c.value = hdr; c.font = WHITE_FONT; c.fill = fl
        c.alignment = CENTER; c.border = THIN_BORDER

    for i, d in enumerate(results):
        dr = hdr_row + 1 + i
        ws.row_dimensions[dr].height = 16
        zebra = GREY_FILL if i % 2 == 0 else None

        ws.cell(row=dr, column=1).value     = i + 1
        ws.cell(row=dr, column=1).font      = BOLD_FONT
        ws.cell(row=dr, column=1).alignment = CENTER
        ws.cell(row=dr, column=1).border    = THIN_BORDER
        if zebra:
            ws.cell(row=dr, column=1).fill = zebra

        style_value(ws.cell(row=dr, column=2), d["Xi"],      fill=zebra)
        style_value(ws.cell(row=dr, column=3), d["YE_i"],    fmt="#,##0",    fill=zebra)
        style_value(ws.cell(row=dr, column=4), d["YU_i"],    fmt="#,##0",    fill=zebra)
        style_value(ws.cell(row=dr, column=5), d["Xi_cv"],   fill=GREEN_FILL)
        style_value(ws.cell(row=dr, column=6), d["ElAppWT"], fill=zebra)
        style_value(ws.cell(row=dr, column=7), d["UrScanWT"],fill=zebra)

    # averages footer
    avg_row = hdr_row + 1 + len(results)
    ws.row_dimensions[avg_row].height = 18
    avg_data = [
        ("AVG", None),
        (np.mean([d["Xi"]       for d in results]), "#,##0.00000"),
        (np.mean([d["YE_i"]     for d in results]), "#,##0.0"),
        (np.mean([d["YU_i"]     for d in results]), "#,##0.0"),
        (np.mean([d["Xi_cv"]    for d in results]), "#,##0.00000"),
        (np.mean([d["ElAppWT"]  for d in results]), "#,##0.00000"),
        (np.mean([d["UrScanWT"] for d in results]), "#,##0.00000"),
    ]
    for ci, (val, fmt) in enumerate(avg_data, start=1):
        c = ws.cell(row=avg_row, column=ci)
        c.border = THIN_BORDER; c.font = BOLD_FONT
        if ci == 1:
            c.value = "AVG"; c.alignment = CENTER
            c.fill = BLUE_FILL; c.font = WHITE_FONT
        elif val is not None:
            c.value = val; c.number_format = fmt; c.alignment = RIGHT
            c.fill = ORANGE_FILL if ci == 5 else LIGHT_FILL

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


# ============================================================
# MAIN
# ============================================================
def main():
    import os
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    for urgent, strategy, rule in DESIGNS:
        input_file = f"{INPUT_DIR}/input-S{strategy}-{urgent}.txt"
        sheet_name = f"S{strategy}-{urgent}-R{rule}"
        print(f"\n{'='*60}")
        print(f"Design: {urgent} urgent slots | Strategy {strategy} | Rule {rule}")
        print(f"{'='*60}")

        sim = Simulation(input_file, TOTAL_WEEKS, R, rule)
        sim.setWeekSchedule()

        X_vals, YE_vals, YU_vals = [], [], []
        results = []

        for r in range(R):
            sim.resetSystem()
            random.seed(r)
            sim.runOneSimulation()

            xi, ye_i, yu_i, el_app, el_scan, ur_scan, ot = compute_replication_outputs(sim)

            X_vals.append(xi)
            YE_vals.append(ye_i)
            YU_vals.append(yu_i)
            results.append({
                "Xi": xi, "YE_i": float(ye_i), "YU_i": float(yu_i),
                "ElAppWT": el_app, "ElScanWT": el_scan,
                "UrScanWT": ur_scan, "OT": ot,
            })

            print(f"  r={r:>3} | OV={xi:.5f} | ElApp={el_app:.3f}"
                  f" | UrScan={ur_scan:.3f} | YE={ye_i} | YU={yu_i}")

        X  = np.array(X_vals,  dtype=float)
        YE = np.array(YE_vals, dtype=float)
        YU = np.array(YU_vals, dtype=float)

        # Known means over post-warmup RUN_WEEKS horizon
        # Elective: Mon-Fri only (5 days), urgent: 4 full days + 2 half days per week
        v_E = 5 * RUN_WEEKS * sim.lambdaElective
        v_U = RUN_WEEKS * (4 * sim.lambdaUrgent[0] + 2 * sim.lambdaUrgent[1])

        # Optimal c = Cov(X, Y) / Var(Y)
        var_YE = float(np.var(YE, ddof=1))
        var_YU = float(np.var(YU, ddof=1))
        c_E = 0.0 if var_YE == 0 else float(np.cov(X, YE, ddof=1)[0, 1] / var_YE)
        c_U = 0.0 if var_YU == 0 else float(np.cov(X, YU, ddof=1)[0, 1] / var_YU)

        # Corrected values: X_cv,i = Xi - c_E*(YE_i - v_E) - c_U*(YU_i - v_U)
        X_cv = X - c_E * (YE - v_E) - c_U * (YU - v_U)

        for i, d in enumerate(results):
            d["Xi_cv"] = float(X_cv[i])

        mean_raw = float(np.mean(X));    std_raw = float(np.std(X,    ddof=1))
        mean_cv  = float(np.mean(X_cv)); std_cv  = float(np.std(X_cv, ddof=1))
        ci_raw   = 1.96 * std_raw / np.sqrt(R)
        ci_cv    = 1.96 * std_cv  / np.sqrt(R)
        reduction = 100 * (1 - (std_cv / std_raw) ** 2) if std_raw > 0 else 0.0

        print(f"\n  v_E={v_E:.1f}  v_U={v_U:.1f}")
        print(f"  c_E={c_E:.6f}  c_U={c_U:.6f}")
        print(f"  Raw : mean={mean_raw:.5f}  std={std_raw:.5f}  CI±{ci_raw:.5f}")
        print(f"  CV  : mean={mean_cv:.5f}  std={std_cv:.5f}  CI±{ci_cv:.5f}")
        print(f"  Variance reduction: {reduction:.2f}%")

        meta = dict(
            urgent=urgent, strategy=strategy, rule=rule,
            v_E=v_E, v_U=v_U, c_E=c_E, c_U=c_U,
            mean_raw=mean_raw, std_raw=std_raw, ci_raw=ci_raw,
            mean_cv=mean_cv,   std_cv=std_cv,   ci_cv=ci_cv,
            reduction_pct=reduction,
        )

        write_design_sheet(wb, sheet_name, results, meta)

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved → {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()